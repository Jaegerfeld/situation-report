# =============================================================================
# Autor:          Robert Seebauer
# Repository:     https://github.com/Jaegerfeld/situation-report
# KI-Unterstützung: Erstellt mit Unterstützung von Claude (Anthropic)
# Erstellt:       03.05.2026
# Geändert:       03.05.2026
# Lizenz:         BSD-3-Clause (siehe LICENSE)
#
# Fachliche Funktion:
#   Acceptance-Tests für helper (JSON Merger). Prüft die fachliche Korrektheit
#   des Merge-Ergebnisses: zusammengeführte JSON-Dateien werden von transform_data
#   verarbeitet und ergeben konsistente Ausgaben. Zwei synthetische JSON-Dateien
#   werden via testdata_generator erzeugt, gemergt und dann transformiert.
# =============================================================================

from __future__ import annotations

import json
from datetime import date
from pathlib import Path

import pytest

from helper.cli import run_merge
from testdata_generator.generator import GeneratorConfig, generate
from testdata_generator.workflow_parser import parse_workflow
from transform_data.transform import run_transform

WORKFLOW_FILE = (
    Path(__file__).parent.parent.parent.parent
    / "testdata_generator"
    / "workflow_ART_A.txt"
)


def _make_json(path: Path, seed: int, count: int = 50) -> Path:
    """Generate count synthetic issues and write them to path."""
    workflow = parse_workflow(WORKFLOW_FILE)
    config = GeneratorConfig(
        project_key="MERGE",
        issue_count=count,
        from_date=date(2025, 1, 1),
        to_date=date(2025, 12, 31),
        completion_rate=0.7,
        seed=seed,
    )
    data = generate(workflow, config)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    return path


@pytest.fixture(scope="module")
def merged_json(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """
    Create two synthetic JSON files (different seeds) and merge them.
    Returns the path to the merged output file.
    """
    tmp = tmp_path_factory.mktemp("helper_acc")
    f1 = _make_json(tmp / "part1.json", seed=1)
    f2 = _make_json(tmp / "part2.json", seed=2)
    out = tmp / "merged.json"
    # deduplicate=False: both files have overlapping ids (generator starts at 100000+i);
    # dedup correctness is verified in unit tests
    run_merge(inputs=[f1, f2], output=out, deduplicate=False)
    return out


@pytest.fixture(scope="module")
def transformed_output(
    merged_json: Path, tmp_path_factory: pytest.TempPathFactory
) -> Path:
    """Run transform_data on the merged JSON and return the output directory."""
    out_dir = tmp_path_factory.mktemp("helper_transform_out")
    run_transform(
        json_file=merged_json,
        workflow_file=WORKFLOW_FILE,
        output_dir=out_dir,
        prefix="MERGE",
    )
    return out_dir


class TestMergedJSONFormat:
    def test_output_is_valid_json(self, merged_json: Path) -> None:
        """The merged file can be loaded as JSON without errors."""
        data = json.loads(merged_json.read_text())
        assert isinstance(data, dict)

    def test_issue_count_is_sum(self, merged_json: Path) -> None:
        """The total in the merged file equals the sum of the two input files."""
        data = json.loads(merged_json.read_text())
        assert data["total"] == 100
        assert len(data["issues"]) == 100

    def test_issues_have_required_fields(self, merged_json: Path) -> None:
        """Every issue in the merged file has key, fields, and changelog."""
        data = json.loads(merged_json.read_text())
        for issue in data["issues"]:
            assert "key" in issue
            assert "fields" in issue
            assert "changelog" in issue

    def test_envelope_is_correct(self, merged_json: Path) -> None:
        """startAt=0, total==maxResults, issues count matches total."""
        data = json.loads(merged_json.read_text())
        assert data["startAt"] == 0
        assert data["total"] == data["maxResults"]
        assert len(data["issues"]) == data["total"]


class TestTransformDataCompatibility:
    def test_transform_data_can_load_merged(self, transformed_output: Path) -> None:
        """transform_data runs without error on the merged JSON."""
        assert transformed_output.exists()

    def test_issue_times_xlsx_created(self, transformed_output: Path) -> None:
        """transform_data produces an IssueTimes XLSX file from the merged JSON."""
        xlsx_files = list(transformed_output.glob("*IssueTimes*.xlsx"))
        assert len(xlsx_files) >= 1

    def test_issue_times_has_rows(self, transformed_output: Path) -> None:
        """The IssueTimes XLSX contains at least one data row."""
        from openpyxl import load_workbook
        xlsx = next(transformed_output.glob("*IssueTimes*.xlsx"))
        wb = load_workbook(xlsx)
        ws = wb.active
        assert ws.max_row > 1
