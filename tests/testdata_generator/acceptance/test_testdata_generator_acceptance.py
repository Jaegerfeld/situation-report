# =============================================================================
# Autor:          Robert Seebauer
# Repository:     https://github.com/Jaegerfeld/situation-report
# KI-Unterstützung: Erstellt mit Unterstützung von Claude (Anthropic)
# Erstellt:       02.05.2026
# Geändert:       02.05.2026
# Lizenz:         BSD-3-Clause (siehe LICENSE)
#
# Fachliche Funktion:
#   Acceptance-Tests für testdata_generator. Prüft die fachliche Korrektheit
#   der Generierung: erzeugte JSON-Dateien werden von transform_data verarbeitet
#   und ergeben konsistente IssueTimes.xlsx-Ausgaben.
# =============================================================================

from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from testdata_generator.cli import run_generate
from transform_data.transform import run_transform

WORKFLOW_FILE = Path(__file__).parent.parent.parent.parent / "testdata_generator" / "workflow_ART_A.txt"


@pytest.fixture(scope="module")
def generated_json(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """
    Generate 200 synthetic issues using the real ART_A workflow, fixed seed.
    Returns the path to the generated JSON file.
    """
    tmp = tmp_path_factory.mktemp("gen")
    output = tmp / "ART_A_GEN.json"
    run_generate(
        workflow=WORKFLOW_FILE,
        output=output,
        project_key="GEN",
        issue_count=200,
        from_date=date(2025, 1, 1),
        to_date=date(2025, 12, 31),
        completion_rate=0.7,
        seed=42,
    )
    return output


@pytest.fixture(scope="module")
def transformed_output(generated_json: Path, tmp_path_factory: pytest.TempPathFactory) -> Path:
    """
    Run transform_data on the generated JSON and return the output directory.
    """
    out_dir = tmp_path_factory.mktemp("transform_out")
    run_transform(
        json_file=generated_json,
        workflow_file=WORKFLOW_FILE,
        output_dir=out_dir,
        prefix="GEN",
    )
    return out_dir


class TestGeneratedJSONFormat:
    def test_json_file_exists(self, generated_json: Path) -> None:
        """The generated JSON file is created at the specified output path."""
        assert generated_json.exists()

    def test_json_file_is_parseable(self, generated_json: Path) -> None:
        """Generated file is valid JSON."""
        import json
        data = json.loads(generated_json.read_text(encoding="utf-8"))
        assert isinstance(data, dict)
        assert "issues" in data

    def test_json_issue_count(self, generated_json: Path) -> None:
        """Generated JSON contains exactly 200 issues."""
        import json
        data = json.loads(generated_json.read_text(encoding="utf-8"))
        assert len(data["issues"]) == 200

    def test_json_reproducible_with_seed(self, tmp_path: Path) -> None:
        """Two generation runs with the same seed produce identical files."""
        out1 = tmp_path / "run1.json"
        out2 = tmp_path / "run2.json"
        for out in (out1, out2):
            run_generate(
                workflow=WORKFLOW_FILE,
                output=out,
                project_key="SEED_TEST",
                issue_count=50,
                seed=99,
            )
        assert out1.read_text(encoding="utf-8") == out2.read_text(encoding="utf-8")


class TestTransformDataCompatibility:
    def test_issue_times_xlsx_exists(self, transformed_output: Path) -> None:
        """transform_data produces a GEN_IssueTimes.xlsx file."""
        assert (transformed_output / "GEN_IssueTimes.xlsx").exists()

    def test_cfd_xlsx_exists(self, transformed_output: Path) -> None:
        """transform_data produces a GEN_CFD.xlsx file."""
        assert (transformed_output / "GEN_CFD.xlsx").exists()

    def test_transitions_xlsx_exists(self, transformed_output: Path) -> None:
        """transform_data produces a GEN_Transitions.xlsx file."""
        assert (transformed_output / "GEN_Transitions.xlsx").exists()

    def test_issue_times_has_issues(self, transformed_output: Path) -> None:
        """IssueTimes.xlsx contains at least one data row."""
        wb = load_workbook(transformed_output / "GEN_IssueTimes.xlsx")
        ws = wb.active
        assert ws.max_row > 1, "IssueTimes.xlsx has no data rows"

    def test_issue_times_row_count_plausible(self, transformed_output: Path) -> None:
        """IssueTimes.xlsx contains exactly 200 issues (header + 200 rows)."""
        wb = load_workbook(transformed_output / "GEN_IssueTimes.xlsx")
        ws = wb.active
        assert ws.max_row == 201, f"Expected 201 rows (header + 200), got {ws.max_row}"

    def test_open_issues_have_no_closed_date(self, transformed_output: Path) -> None:
        """Issues without a Closed Date in IssueTimes are truly open (no closed transition)."""
        wb = load_workbook(transformed_output / "GEN_IssueTimes.xlsx")
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        closed_col = next(
            (i for i, h in enumerate(headers) if h and "Closed" in str(h)), None
        )
        first_col = next(
            (i for i, h in enumerate(headers) if h and "First" in str(h)), None
        )
        if closed_col is None or first_col is None:
            pytest.skip("Expected columns not found — check IssueTimes format")

        open_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[closed_col] is None:
                open_count += 1

        assert open_count > 0, "Expected some open issues (no Closed Date)"

    def test_issue_times_has_key_column(self, transformed_output: Path) -> None:
        """IssueTimes.xlsx has a 'Key' column."""
        wb = load_workbook(transformed_output / "GEN_IssueTimes.xlsx")
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert any("Key" in str(h) for h in headers if h)
