# =============================================================================
# Autor:          Robert Seebauer
# Repository:     https://github.com/Jaegerfeld/situation-report
# KI-Unterstützung: Erstellt mit Unterstützung von Claude (Anthropic)
# Erstellt:       02.09.2026
# Geändert:       02.09.2026
# Lizenz:         BSD-3-Clause (siehe LICENSE)
#
# Fachliche Funktion:
#   Unit-Tests für das Portfolio-Szenario (testdata_generator/scenario.py):
#   Artefakt-Vollständigkeit, Roundtrip durch die echten Parser
#   (load_solution_config, load_pi_config, Pooling), die eingebauten
#   Demo-Geschichten (Ausreißer, schwache Quelle, stage_map) und
#   Determinismus bei festem Seed und Referenzdatum.
# =============================================================================

from __future__ import annotations

from datetime import date, timedelta

import pytest

from build_reports.pi_config import load_pi_config
from portfolio.aggregator import build_pooled_report_data
from portfolio.solution_config import KIND_PORTFOLIO, load_solution_config
from portfolio.summary import CONFIDENCE_LOW, SourceQuality
from testdata_generator.scenario import build_portfolio_scenario

REF = date(2025, 6, 30)


@pytest.fixture(scope="module")
def scenario(tmp_path_factory):
    """Build the scenario once per module (generation takes a few seconds)."""
    out = tmp_path_factory.mktemp("demo")
    paths = build_portfolio_scenario(out, seed=42, reference=REF,
                                     log=lambda m: None)
    return out, paths


class TestArtifacts:
    def test_all_expected_files_exist(self, scenario) -> None:
        out, paths = scenario
        for key in ("portfolio", "solution_alpha", "solution_beta",
                    "risks_alpha", "risks_beta", "nfr_alpha", "nfr_beta",
                    "capabilities_alpha", "capabilities_beta",
                    "dependencies_alpha", "dependencies_beta",
                    "decisions_alpha", "decisions_beta",
                    "snapshot_prev", "snapshot_now",
                    "slo_alpha", "slo_beta", "dora_alpha", "dora_beta",
                    "flow_alpha", "flow_beta",
                    "themes_alpha", "themes_beta",
                    "pi_config", "readme"):
            assert paths[key].exists(), key
        assert len(list((out / "arts").glob("*_IssueTimes.xlsx"))) == 6
        assert len(list((out / "arts").glob("*_Transitions.xlsx"))) == 6
        # Die schwache Quelle Beta-3 liefert bewusst kein CFD.
        assert len(list((out / "arts").glob("*_CFD.xlsx"))) == 5
        assert len(list((out / "raw").glob("*_jira.json"))) == 6

    def test_pi_config_loads_and_covers_window(self, scenario) -> None:
        _, paths = scenario
        intervals = load_pi_config(paths["pi_config"])
        assert len(intervals) >= 4
        assert intervals[-1].to_date == REF


class TestConfigsRoundtrip:
    def test_solution_beta_carries_stage_map(self, scenario) -> None:
        _, paths = scenario
        cfg = load_solution_config(paths["solution_beta"])
        assert cfg.stage_map is not None
        assert list(cfg.stage_map.stages.keys()) == ["Vorlauf", "Umsetzung", "Fertig"]
        assert cfg.stage_map.first_stage == "Umsetzung"

    def test_solution_alpha_has_no_stage_map(self, scenario) -> None:
        _, paths = scenario
        assert load_solution_config(paths["solution_alpha"]).stage_map is None

    def test_portfolio_references_both_solutions(self, scenario) -> None:
        _, paths = scenario
        cfg = load_solution_config(paths["portfolio"])
        assert cfg.kind == KIND_PORTFOLIO
        assert len(cfg.members) == 2


class TestBuiltInStories:
    def test_pooling_works_and_beta_uses_custom_stages(self, scenario) -> None:
        _, paths = scenario
        cfg = load_solution_config(paths["solution_beta"])
        pooled = build_pooled_report_data(cfg, log=lambda m: None)
        assert pooled.stages == ["Vorlauf", "Umsetzung", "Fertig"]
        assert len(pooled.issues) > 0

    def test_weak_source_is_low_confidence(self, scenario) -> None:
        _, paths = scenario
        cfg = load_solution_config(paths["solution_beta"])
        qualities: list[SourceQuality] = []
        build_pooled_report_data(cfg, log=lambda m: None, quality_sink=qualities)
        by_name = {q.label: q for q in qualities}
        weak = by_name["ART Beta-3"]
        assert weak.confidence == CONFIDENCE_LOW
        assert weak.has_cfd is False
        # Datenstand der schwachen Quelle liegt ~60 Tage zurück.
        assert weak.age_days is not None and weak.age_days >= 55

    def test_healthy_sources_have_cfd_and_fresh_data(self, scenario) -> None:
        _, paths = scenario
        cfg = load_solution_config(paths["solution_alpha"])
        qualities: list[SourceQuality] = []
        pooled = build_pooled_report_data(cfg, log=lambda m: None,
                                          quality_sink=qualities)
        assert all(q.has_cfd for q in qualities)
        # Frisch relativ zum Szenario-Referenzdatum (nicht zu heute) pruefen:
        for q in qualities:
            assert q.data_as_of is not None
            assert (REF - q.data_as_of).days <= 30
        assert len(pooled.issues) == 360

    def test_roam_registers_load_and_carry_two_aging_risks(self, scenario) -> None:
        from portfolio.risks_config import ROAM_OWNED, load_risks
        _, paths = scenario
        alpha = load_risks(paths["risks_alpha"])
        beta = load_risks(paths["risks_beta"])
        assert alpha.risks and beta.risks
        aging = [r for r in alpha.risks + beta.risks
                 if r.roam == ROAM_OWNED and r.status_since is not None
                 and (REF - r.status_since).days > 30]
        assert len(aging) == 2
        assert load_solution_config(paths["solution_alpha"]).risks
        assert load_solution_config(paths["solution_beta"]).risks

    def test_portfolio_collects_risks_from_both_solutions(self, scenario) -> None:
        from portfolio.aggregator import _collect_risks
        _, paths = scenario
        cfg = load_solution_config(paths["portfolio"])
        entries = _collect_risks(cfg, log=lambda m: None)
        assert {source for source, _ in entries} == {"Solution Alpha", "Solution Beta"}
        assert len(entries) == 9

    def test_nfr_registers_load_and_tell_their_story(self, scenario) -> None:
        from portfolio.nfr_config import (
            RUNWAY_GAP,
            STATUS_VIOLATED,
            load_nfr,
        )
        _, paths = scenario
        alpha = load_nfr(paths["nfr_alpha"])
        beta = load_nfr(paths["nfr_beta"])
        assert alpha.nfrs and alpha.runway and beta.nfrs and beta.runway
        # Die Geschichte: genau eine verletzte NFR und eine überfällige Lücke,
        # beide bei Solution Beta.
        assert not any(n.status == STATUS_VIOLATED for n in alpha.nfrs)
        violated = [n for n in beta.nfrs if n.status == STATUS_VIOLATED]
        assert len(violated) == 1
        gaps = [r for r in beta.runway if r.status == RUNWAY_GAP]
        assert len(gaps) == 1
        assert gaps[0].needed_by is not None and gaps[0].needed_by < REF

    def test_portfolio_collects_nfr_from_both_solutions(self, scenario) -> None:
        from portfolio.aggregator import _collect_nfr
        _, paths = scenario
        cfg = load_solution_config(paths["portfolio"])
        nfrs, runway = _collect_nfr(cfg, log=lambda m: None)
        assert {source for source, _ in nfrs} == {"Solution Alpha", "Solution Beta"}
        assert len(nfrs) == 6
        assert len(runway) == 4

    def test_capability_maps_load_and_tell_their_story(self, scenario) -> None:
        from portfolio.capability_config import HEALTH_CRITICAL, load_capabilities
        _, paths = scenario
        alpha = load_capabilities(paths["capabilities_alpha"])
        beta = load_capabilities(paths["capabilities_beta"])
        assert alpha.capabilities and beta.capabilities
        # Die Geschichte: genau eine kritische Capability (Beta, schwache
        # Quelle) und genau eine uncovered Capability (Alpha, kein ART).
        assert not any(c.health == HEALTH_CRITICAL for c in alpha.capabilities)
        critical = [c for c in beta.capabilities if c.health == HEALTH_CRITICAL]
        assert len(critical) == 1
        assert "ART Beta-3" in critical[0].arts
        uncovered = [c for c in alpha.capabilities + beta.capabilities
                     if not c.arts]
        assert len(uncovered) == 1

    def test_portfolio_collects_capabilities_without_warnings(self, scenario) -> None:
        from portfolio.aggregator import _collect_capabilities
        _, paths = scenario
        cfg = load_solution_config(paths["portfolio"])
        warnings: list[str] = []
        entries = _collect_capabilities(cfg, log=warnings.append)
        assert {source for source, _ in entries} == {"Solution Alpha", "Solution Beta"}
        assert len(entries) == 6
        # Alle gemappten ART-Namen existieren als Member — keine Drift-Warnung.
        assert warnings == []

    def test_dependency_registers_load_and_tell_their_story(self, scenario) -> None:
        from portfolio.dependency_config import (
            DEP_BLOCKED,
            load_dependencies,
        )
        _, paths = scenario
        alpha = load_dependencies(paths["dependencies_alpha"])
        beta = load_dependencies(paths["dependencies_beta"])
        assert alpha.dependencies and beta.dependencies
        # Die Geschichte: genau eine blockierte, überfällige Abhängigkeit
        # (Alpha-1 braucht den Ausreißer Alpha-3) ...
        blocked = [d for d in alpha.dependencies if d.status == DEP_BLOCKED]
        assert len(blocked) == 1
        assert blocked[0].to_art == "ART Alpha-3"
        assert blocked[0].due is not None and blocked[0].due < REF
        # ... und eine Cross-Solution-Integration (Beta braucht Alpha).
        cross = [d for d in beta.dependencies if d.to_art == "ART Alpha-1"]
        assert len(cross) == 1

    def test_portfolio_collects_dependencies_from_both_solutions(self, scenario) -> None:
        from portfolio.aggregator import _collect_dependencies
        _, paths = scenario
        cfg = load_solution_config(paths["portfolio"])
        entries = _collect_dependencies(cfg, log=lambda m: None)
        assert {source for source, _ in entries} == {"Solution Alpha", "Solution Beta"}
        assert len(entries) == 5

    def test_decision_logs_load_and_tell_their_story(self, scenario) -> None:
        from portfolio.decision_config import (
            ASSUMPTION_OPEN,
            KIND_ASSUMPTION,
            load_decisions,
        )
        _, paths = scenario
        alpha = load_decisions(paths["decisions_alpha"])
        beta = load_decisions(paths["decisions_beta"])
        assert alpha.entries and beta.entries
        # Die Geschichte: genau eine offene Annahme mit überschrittenem
        # Prüfdatum (Beta) — und Alphas supersedes-Kette ist intakt.
        stale = [e for e in alpha.entries + beta.entries
                 if e.kind == KIND_ASSUMPTION and e.status == ASSUMPTION_OPEN
                 and e.review_by is not None and e.review_by < REF]
        assert len(stale) == 1
        assert stale[0].entry_id == "AS-B1"
        chained = [e for e in alpha.entries if e.supersedes]
        assert len(chained) == 1
        assert chained[0].supersedes in {e.entry_id for e in alpha.entries}

    def test_portfolio_collects_decisions_from_both_solutions(self, scenario) -> None:
        from portfolio.aggregator import _collect_decisions
        _, paths = scenario
        cfg = load_solution_config(paths["portfolio"])
        entries = _collect_decisions(cfg, log=lambda m: None)
        assert {source for source, _ in entries} == {"Solution Alpha", "Solution Beta"}
        assert len(entries) == 5

    def test_delta_briefing_snapshots_tell_their_story(self, scenario) -> None:
        from portfolio.delta import compute_delta
        from portfolio.snapshot import load_snapshot
        _, paths = scenario
        prev = load_snapshot(paths["snapshot_prev"])
        now = load_snapshot(paths["snapshot_now"])
        assert prev.as_of == REF - timedelta(days=14)
        assert now.as_of == REF

        delta = compute_delta(prev, now)
        assert not delta.quiet
        assert delta.period_days == 14
        # Durchsatz: im Zeitraum wurde sichtbar fertiggestellt.
        assert delta.completed_delta > 0
        # Die Geschichte: Beta-3-Konfidenz verfällt medium → low, ...
        assert [(c.entry_id, c.fields["confidence"])
                for c in delta.confidence_changes] \
            == [("ART Beta-3", ("medium", "low"))]
        # ... AD-1 eskaliert at_risk → blocked, ...
        deps = delta.governance["dependencies"]
        assert [(c.entry_id, c.fields["status"]) for c in deps.changed] \
            == [("AD-1", ("at_risk", "blocked"))]
        # ... Risiko BR-2 ist neu, ...
        assert [e["id"] for e in delta.governance["risks"].added] == ["BR-2"]
        # ... und Runway-Lücke + offene Annahme sind frisch überfällig.
        assert [e["id"] for e in delta.governance["runway"].newly_overdue] \
            == ["BRW-1"]
        assert [e["id"] for e in delta.governance["decisions"].newly_overdue] \
            == ["AS-B1"]

    def test_slo_and_dora_registers_tell_their_story(self, scenario) -> None:
        from portfolio.aggregator import _collect_delivery, _collect_slo
        from portfolio.dora_config import unit_tier
        from portfolio.slo_config import slo_status
        _, paths = scenario
        cfg = load_solution_config(paths["portfolio"])
        slo_entries = _collect_slo(cfg, log=lambda m: None)
        assert len(slo_entries) == 5
        # Die Geschichte: Betas Sync-API reisst ihr SLO, Alphas Checkout
        # hat kaum noch Budget.
        by_status = {r.service: slo_status(r) for _, r in slo_entries}
        assert by_status["Order Sync API"] == "breached"
        assert by_status["Checkout"] == "at_risk"

        dora_entries, quality_entries = _collect_delivery(cfg, log=lambda m: None)
        assert len(dora_entries) == 5 and len(quality_entries) == 3
        tiers = {r.unit: unit_tier(r) for _, r in dora_entries}
        # Beta-3 ist auch im Delivery-Bild low; Alpha-1 liefert elite.
        assert tiers["ART Beta-3"] == "low"
        assert tiers["ART Alpha-1"] == "elite"
        weak = next(q for _, q in quality_entries if q.unit == "ART Beta-3")
        assert weak.maintainability == "D" and weak.critical_issues == 7

    def test_flow_problem_backlog_tells_the_workshop_story(self, scenario) -> None:
        from portfolio.aggregator import _collect_flow_problems
        _, paths = scenario
        cfg = load_solution_config(paths["portfolio"])
        entries = _collect_flow_problems(cfg, log=lambda m: None)
        assert len(entries) == 4
        # Das Workshop-Muster: FP-A1 und FP-B1 überleben Konferenzen.
        survivors = {p.problem_id for _, p in entries if p.survived}
        assert survivors == {"FP-A1", "FP-B1"}
        fp_b1 = next(p for _, p in entries if p.problem_id == "FP-B1")
        assert fp_b1.cross_vs and fp_b1.conferences == 4

    def test_conference_preread_bundles_the_inputs(self, scenario) -> None:
        from portfolio.aggregator import render_conference_html
        _, paths = scenario
        cfg = load_solution_config(paths["portfolio"])
        conf = render_conference_html(cfg, conference_date=REF,
                                      log=lambda m: None)
        for marker in ("Value-Stream Conference", "Input 1", "Input 2",
                       "Input 3", "Flow-Problem Backlog"):
            assert marker in conf
        # Impediment-Backlog vor den verwandten Governance-Sichten.
        assert conf.index("Flow-Problem Backlog") < conf.index("ROAM")

    def test_themes_register_tells_orphan_and_zombie_story(self, scenario) -> None:
        from portfolio.aggregator import _collect_themes
        from portfolio.themes_config import ThemesRegister, orphan_theme_ids, zombie_epics
        _, paths = scenario
        cfg = load_solution_config(paths["portfolio"])
        theme_entries, epic_entries = _collect_themes(cfg, log=lambda m: None)
        assert len(theme_entries) == 3 and len(epic_entries) == 7
        merged = ThemesRegister(
            themes=[t for _, t in theme_entries],
            epics=[e for _, e in epic_entries])
        assert orphan_theme_ids(merged) == {"T-A2"}
        assert [e.epic_id for e in zombie_epics(merged)] == ["EP-A9"]

    def test_delta_documents_updated_roadmaps(self, scenario) -> None:
        from portfolio.delta import compute_delta
        from portfolio.snapshot import load_snapshot
        _, paths = scenario
        delta = compute_delta(load_snapshot(paths["snapshot_prev"]),
                              load_snapshot(paths["snapshot_now"]))
        by_id = {c.entry_id: c for c in delta.governance["epics"].changed}
        # EP-A9 verlor sein Theme (Zombie, worsened); EP-B2 wurde P2 -> P1.
        assert by_id["EP-A9"].fields["theme"] == ("T-A1", "") \
            and by_id["EP-A9"].worsened
        assert by_id["EP-B2"].fields["horizon"] == ("P2", "P1")

    def test_outlier_art_has_clearly_higher_cycle_time(self, scenario) -> None:
        from portfolio.aggregator import load_members
        from portfolio.summary import compute_summary
        _, paths = scenario
        cfg = load_solution_config(paths["solution_alpha"])
        summaries = {d.source_prefix: compute_summary(d, d.source_prefix)
                     for d in load_members(cfg, log=lambda m: None)}
        outlier = summaries["ART Alpha-3"].median_ct
        others = [summaries["ART Alpha-1"].median_ct, summaries["ART Alpha-2"].median_ct]
        assert outlier is not None and all(o is not None for o in others)
        assert outlier > 1.5 * max(others)


class TestDeterminism:
    def test_same_seed_same_issue_times(self, tmp_path) -> None:
        import openpyxl
        a = tmp_path / "a"
        b = tmp_path / "b"
        build_portfolio_scenario(a, seed=7, reference=REF, log=lambda m: None)
        build_portfolio_scenario(b, seed=7, reference=REF, log=lambda m: None)
        wa = openpyxl.load_workbook(a / "arts" / "Alpha-1_IssueTimes.xlsx").active
        wb = openpyxl.load_workbook(b / "arts" / "Alpha-1_IssueTimes.xlsx").active
        rows_a = [tuple(r) for r in wa.iter_rows(values_only=True)]
        rows_b = [tuple(r) for r in wb.iter_rows(values_only=True)]
        assert rows_a == rows_b
