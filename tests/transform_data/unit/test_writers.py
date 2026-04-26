# =============================================================================
# Autor:          Robert Seebauer
# Repository:     https://github.com/Jaegerfeld/situation-report
# KI-Unterstützung: Erstellt mit Unterstützung von Claude (Anthropic)
# Erstellt:       22.04.2026
# Geändert:       26.04.2026
# Lizenz:         BSD-3-Clause (siehe LICENSE)
#
# Fachliche Funktion:
#   Unit-Tests für transform_data.writers. Prüft write_cfd (tägliche
#   Eintrittszählungen), write_transitions (chronologische Statuswechsel)
#   und write_issue_times (Verweildauer je Stage pro Issue).
# =============================================================================

"""Unit tests for transform_data.writers."""

from datetime import date, datetime, timezone, timedelta
from pathlib import Path

import openpyxl
import pytest

from transform_data.processor import IssueRecord, Transition
from transform_data.workflow import parse_workflow
from transform_data.writers import write_cfd, write_transitions, write_issue_times


WORKFLOW_TEXT = "\n".join([
    "Funnel",
    "Analysis",
    "Done",
    "<First>Analysis",
    "<Closed>Done",
])

_TZ = timezone(timedelta(hours=1))


def _dt(d: date, h: int = 9) -> datetime:
    """Return a datetime at the given date and hour in the test timezone."""
    return datetime(d.year, d.month, d.day, h, 0, 0, tzinfo=_TZ)


def _record(
    key: str,
    created: datetime,
    initial_stage: str,
    transitions: list[tuple[str, datetime]],
) -> IssueRecord:
    """Build an IssueRecord with empty stage_minutes (for write_cfd / write_transitions tests).

    Args:
        key:           Issue key string.
        created:       Issue creation datetime.
        initial_stage: Stage the issue was in before its first transition.
        transitions:   List of (stage_label, timestamp) tuples after creation.

    Returns:
        IssueRecord with stage_minutes={} and None dates.
    """
    trans = [Transition(key=key, label="Created", timestamp=created)]
    for label, ts in transitions:
        trans.append(Transition(key=key, label=label, timestamp=ts))
    return IssueRecord(
        project="TEST", key=key, issuetype="Story", status="Open",
        created=created, component="", first_date=None, inprogress_date=None,
        closed_date=None, stage_minutes={}, resolution="",
        initial_stage=initial_stage, transitions=trans,
    )


def _full_record(
    key: str,
    created: datetime,
    stage_minutes: dict[str, int],
    first_date: datetime | None = None,
    closed_date: datetime | None = None,
    resolution: str = "",
) -> IssueRecord:
    """Build an IssueRecord with stage_minutes populated (for write_issue_times tests).

    Args:
        key:           Issue key string.
        created:       Issue creation datetime.
        stage_minutes: Dict mapping stage name to minutes spent.
        first_date:    Optional first-date milestone timestamp.
        closed_date:   Optional closed-date milestone timestamp.
        resolution:    Resolution string (default empty).

    Returns:
        IssueRecord suitable for write_issue_times.
    """
    trans = [Transition(key=key, label="Created", timestamp=created)]
    return IssueRecord(
        project="TEST", key=key, issuetype="Story", status="Open",
        created=created, component="Comp-A",
        first_date=first_date, inprogress_date=None, closed_date=closed_date,
        stage_minutes=stage_minutes, resolution=resolution,
        initial_stage=None, transitions=trans,
    )


@pytest.fixture
def workflow(tmp_path):
    """Parse a three-stage workflow (Funnel → Analysis → Done)."""
    wf_file = tmp_path / "workflow.txt"
    wf_file.write_text(WORKFLOW_TEXT)
    return parse_workflow(wf_file)


def _read_cfd(path: Path) -> dict[str, dict[str, int]]:
    """Read CFD.xlsx and return {date_str: {stage: count}}."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        day = row[0]
        result[day] = {headers[i]: row[i] for i in range(1, len(headers))}
    return result


def _read_rows(path: Path) -> tuple[list[str], list[list]]:
    """Read an XLSX and return (header, data_rows)."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    data = [list(r) for r in rows[1:]]
    return header, data


# ---------------------------------------------------------------------------
# write_cfd
# ---------------------------------------------------------------------------

class TestWriteCfd:
    """Tests for write_cfd — daily entry count semantics."""

    def test_initial_stage_counted_on_creation_date(self, workflow, tmp_path):
        """Issue with no transitions: initial_stage counted on creation day."""
        day = date(2025, 1, 1)
        record = _record("A-1", _dt(day), "Funnel", [])
        ref = _dt(date(2025, 1, 2))
        out = tmp_path / "CFD.xlsx"

        write_cfd([record], workflow, out, ref)
        data = _read_cfd(out)

        assert data["01.01.2025"]["Funnel"] == 1
        assert data["01.01.2025"]["Analysis"] == 0
        assert data["01.01.2025"]["Done"] == 0

    def test_transition_counted_on_transition_date(self, workflow, tmp_path):
        """Transition to Analysis on Jan 2 is counted on that day, not creation day."""
        create_day = date(2025, 1, 1)
        trans_day = date(2025, 1, 2)
        record = _record("A-1", _dt(create_day), "Funnel", [
            ("Analysis", _dt(trans_day)),
        ])
        ref = _dt(date(2025, 1, 3))
        out = tmp_path / "CFD.xlsx"

        write_cfd([record], workflow, out, ref)
        data = _read_cfd(out)

        assert data["01.01.2025"]["Funnel"] == 1
        assert data["01.01.2025"]["Analysis"] == 0
        assert data["02.01.2025"]["Analysis"] == 1
        assert data["02.01.2025"]["Funnel"] == 0

    def test_two_issues_same_day(self, workflow, tmp_path):
        """Two issues entering the same stage on the same day are summed."""
        day = date(2025, 1, 1)
        r1 = _record("A-1", _dt(day), "Funnel", [])
        r2 = _record("A-2", _dt(day), "Funnel", [])
        out = tmp_path / "CFD.xlsx"

        write_cfd([r1, r2], workflow, out, _dt(date(2025, 1, 1)))
        data = _read_cfd(out)

        assert data["01.01.2025"]["Funnel"] == 2

    def test_unmapped_label_not_counted(self, workflow, tmp_path):
        """Transitions with unmapped stage labels are not counted in any stage column."""
        day = date(2025, 1, 1)
        trans_day = date(2025, 1, 2)
        record = _record("A-1", _dt(day), "Funnel", [
            ("Unknown Status", _dt(trans_day)),
        ])
        out = tmp_path / "CFD.xlsx"

        write_cfd([record], workflow, out, _dt(date(2025, 1, 2)))
        data = _read_cfd(out)

        assert data["02.01.2025"]["Funnel"] == 0
        assert data["02.01.2025"]["Analysis"] == 0
        assert data["02.01.2025"]["Done"] == 0

    def test_days_with_no_entries_are_zero(self, workflow, tmp_path):
        """Days without any transitions have zero counts for all stages."""
        record = _record("A-1", _dt(date(2025, 1, 1)), "Funnel", [
            ("Done", _dt(date(2025, 1, 5))),
        ])
        out = tmp_path / "CFD.xlsx"

        write_cfd([record], workflow, out, _dt(date(2025, 1, 5)))
        data = _read_cfd(out)

        for d in ["02.01.2025", "03.01.2025", "04.01.2025"]:
            assert all(v == 0 for v in data[d].values())

    def test_empty_records_writes_no_file(self, workflow, tmp_path):
        """write_cfd with empty records does not create a file."""
        out = tmp_path / "CFD.xlsx"
        write_cfd([], workflow, out, _dt(date(2025, 1, 1)))
        assert not out.exists()


# ---------------------------------------------------------------------------
# write_transitions
# ---------------------------------------------------------------------------

class TestWriteTransitions:
    """Tests for write_transitions — chronological status history per issue."""

    def test_header_columns(self, tmp_path):
        """Output file has exactly Key, Transition, Timestamp as header."""
        out = tmp_path / "T.xlsx"
        write_transitions([], out)
        header, _ = _read_rows(out)
        assert header == ["Key", "Transition", "Timestamp"]

    def test_one_row_per_transition(self, tmp_path):
        """Issue with Created + one transition produces two data rows."""
        record = _record("T-1", _dt(date(2025, 1, 1)), "Funnel", [
            ("Analysis", _dt(date(2025, 1, 2))),
        ])
        out = tmp_path / "T.xlsx"
        write_transitions([record], out)
        _, data = _read_rows(out)
        assert len(data) == 2  # Created + Analysis

    def test_timestamp_format(self, tmp_path):
        """Timestamp column is formatted as DD.MM.YYYY HH:MM:SS."""
        created = datetime(2025, 3, 15, 14, 30, 0, tzinfo=_TZ)
        record = _record("T-1", created, "Funnel", [])
        out = tmp_path / "T.xlsx"
        write_transitions([record], out)
        _, data = _read_rows(out)
        assert data[0][2] == "15.03.2025 14:30:00"

    def test_key_and_label_written(self, tmp_path):
        """Key and transition label are written correctly."""
        record = _record("T-42", _dt(date(2025, 1, 1)), "Funnel", [
            ("Done", _dt(date(2025, 1, 3))),
        ])
        out = tmp_path / "T.xlsx"
        write_transitions([record], out)
        _, data = _read_rows(out)
        assert data[0][0] == "T-42"
        assert data[0][1] == "Created"
        assert data[1][0] == "T-42"
        assert data[1][1] == "Done"

    def test_multiple_issues_all_rows_written(self, tmp_path):
        """Rows from all issues appear: 1 transition for T-1, 2 for T-2 → 3 rows."""
        d = date(2025, 1, 1)
        r1 = _record("T-1", _dt(d), "Funnel", [])
        r2 = _record("T-2", _dt(d), "Funnel", [
            ("Analysis", _dt(date(2025, 1, 2))),
        ])
        out = tmp_path / "T.xlsx"
        write_transitions([r1, r2], out)
        _, data = _read_rows(out)
        assert len(data) == 3


# ---------------------------------------------------------------------------
# write_issue_times
# ---------------------------------------------------------------------------

class TestWriteIssueTimes:
    """Tests for write_issue_times — per-issue stage durations."""

    def test_header_contains_fixed_and_stage_columns(self, workflow, tmp_path):
        """Header includes fixed columns and all workflow stage names."""
        out = tmp_path / "IT.xlsx"
        write_issue_times([], workflow, out)
        header, _ = _read_rows(out)
        for col in ("Project", "Key", "Issuetype", "Status", "Created Date",
                    "First Date", "Closed Date", "Resolution"):
            assert col in header
        for stage in ("Funnel", "Analysis", "Done"):
            assert stage in header

    def test_one_row_per_issue(self, workflow, tmp_path):
        """Two issues produce two data rows."""
        d = _dt(date(2025, 1, 1))
        records = [
            _full_record("T-1", d, {"Funnel": 60, "Analysis": 30, "Done": 0}),
            _full_record("T-2", d, {"Funnel": 0, "Analysis": 120, "Done": 60}),
        ]
        out = tmp_path / "IT.xlsx"
        write_issue_times(records, workflow, out)
        _, data = _read_rows(out)
        assert len(data) == 2

    def test_stage_minutes_written_correctly(self, workflow, tmp_path):
        """Stage minute values from the record appear in the correct columns."""
        d = _dt(date(2025, 1, 1))
        record = _full_record("T-1", d, {"Funnel": 90, "Analysis": 45, "Done": 15})
        out = tmp_path / "IT.xlsx"
        write_issue_times([record], workflow, out)
        header, data = _read_rows(out)
        row = dict(zip(header, data[0]))
        assert row["Funnel"] == 90
        assert row["Analysis"] == 45
        assert row["Done"] == 15

    def test_none_dates_produce_empty_or_none(self, workflow, tmp_path):
        """None first_date and closed_date are written as empty (read back as None or '')."""
        d = _dt(date(2025, 1, 1))
        record = _full_record("T-1", d, {"Funnel": 0, "Analysis": 0, "Done": 0})
        out = tmp_path / "IT.xlsx"
        write_issue_times([record], workflow, out)
        header, data = _read_rows(out)
        row = dict(zip(header, data[0]))
        assert row["First Date"] in (None, "")
        assert row["Closed Date"] in (None, "")

    def test_dates_formatted_as_string(self, workflow, tmp_path):
        """Non-None dates are written as DD.MM.YYYY HH:MM:SS strings."""
        created = datetime(2025, 6, 1, 8, 0, 0, tzinfo=_TZ)
        first = datetime(2025, 6, 3, 10, 0, 0, tzinfo=_TZ)
        record = _full_record("T-1", created, {"Funnel": 0, "Analysis": 0, "Done": 0},
                              first_date=first)
        out = tmp_path / "IT.xlsx"
        write_issue_times([record], workflow, out)
        header, data = _read_rows(out)
        row = dict(zip(header, data[0]))
        assert row["Created Date"] == "01.06.2025 08:00:00"
        assert row["First Date"] == "03.06.2025 10:00:00"
