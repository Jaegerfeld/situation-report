# =============================================================================
# Autor:          Robert Seebauer
# Repository:     https://github.com/Jaegerfeld/situation-report
# KI-Unterstützung: Erstellt mit Unterstützung von Claude (Anthropic)
# Erstellt:       15.04.2026
# Geändert:       18.04.2026
# Lizenz:         BSD-3-Clause (siehe LICENSE)
#
# Fachliche Funktion:
#   Unit-Tests für export.py. pio.write_image wird gemockt, da kaleido einen
#   externen Browser-Prozess benötigt. Geprüft wird die Steuerlogik: korrekte
#   Dateinamen, Verzeichniserstellung, Fehlerbehandlung bei leerer Liste sowie
#   das Fallback-Verhalten für mehrere Figures ohne pypdf.
# =============================================================================

from pathlib import Path
from unittest.mock import MagicMock, call, patch

import plotly.graph_objects as go
import pytest

from build_reports.export import export_pdf, export_png, write_report_excel, write_zero_day_excel
from build_reports.loader import IssueRecord


@pytest.fixture
def simple_figure() -> go.Figure:
    """Return a basic plotly Scatter figure for export tests."""
    return go.Figure(go.Scatter(x=[1, 2, 3], y=[1, 4, 9]))


@pytest.fixture
def two_figures(simple_figure) -> list[go.Figure]:
    """Return two distinct plotly figures for multi-page export tests."""
    return [simple_figure, go.Figure(go.Bar(x=["a"], y=[1]))]


class TestExportPdf:
    """Tests for export_pdf() — PDF generation and routing logic."""

    def test_raises_on_empty_list(self, tmp_path):
        """Passing an empty figures list raises ValueError."""
        with pytest.raises(ValueError, match="No figures"):
            export_pdf([], tmp_path / "out.pdf")

    def test_single_figure_calls_write_image(self, tmp_path, simple_figure):
        """A single figure is written directly via pio.write_image with PDF format."""
        out = tmp_path / "report.pdf"
        with patch("build_reports.export.pio.write_image") as mock_write:
            export_pdf([simple_figure], out)
        mock_write.assert_called_once_with(
            simple_figure, str(out), format="pdf", width=1400, height=700
        )

    def test_creates_parent_dirs(self, tmp_path, simple_figure):
        """export_pdf creates intermediate parent directories if they do not exist."""
        out = tmp_path / "subdir" / "nested" / "report.pdf"
        with patch("build_reports.export.pio.write_image"):
            export_pdf([simple_figure], out)
        assert out.parent.exists()

    def test_multiple_figures_fallback_without_pypdf(self, tmp_path, two_figures):
        """Without pypdf, separate numbered PDFs are created for each figure."""
        out = tmp_path / "multi.pdf"
        with patch("build_reports.export.pio.write_image") as mock_write, \
             patch.dict("sys.modules", {"pypdf": None}):
            export_pdf(two_figures, out)
        assert mock_write.call_count == 2
        called_paths = [Path(c.args[1]) for c in mock_write.call_args_list]
        assert any("page1" in p.name for p in called_paths)
        assert any("page2" in p.name for p in called_paths)

    def test_custom_dimensions(self, tmp_path, simple_figure):
        """Custom width/height parameters are forwarded to pio.write_image."""
        out = tmp_path / "report.pdf"
        with patch("build_reports.export.pio.write_image") as mock_write:
            export_pdf([simple_figure], out, width=800, height=600)
        mock_write.assert_called_once_with(
            simple_figure, str(out), format="pdf", width=800, height=600
        )


def _make_record(key: str, project: str = "ART") -> IssueRecord:
    """Create a minimal IssueRecord suitable for use as a zero-day issue record.

    Args:
        key:     Unique issue key.
        project: Project key (default 'ART').

    Returns:
        IssueRecord with first_date = closed_date = 2025-01-05 (zero-day scenario).
    """
    from datetime import datetime
    return IssueRecord(
        project=project, key=key, issuetype="Feature", status="Done",
        created=datetime(2025, 1, 1), component="",
        first_date=datetime(2025, 1, 5), implementation_date=None,
        closed_date=datetime(2025, 1, 5),
        stage_minutes={}, resolution="Done",
    )


class TestWriteZeroDayExcel:
    """Tests for write_zero_day_excel() — zero-day issue XLSX export."""

    def test_creates_file(self, tmp_path):
        """An XLSX file is created at the given path."""
        path = tmp_path / "zero_day.xlsx"
        write_zero_day_excel([_make_record("A-1")], path)
        assert path.exists()

    def test_creates_parent_dirs(self, tmp_path):
        """Intermediate parent directories are created if they do not exist."""
        path = tmp_path / "sub" / "zero_day.xlsx"
        write_zero_day_excel([_make_record("A-1")], path)
        assert path.exists()

    def test_file_contains_header_and_data(self, tmp_path):
        """The exported file starts with a header row followed by one row per record."""
        from openpyxl import load_workbook
        path = tmp_path / "zero_day.xlsx"
        write_zero_day_excel([_make_record("A-1"), _make_record("A-2")], path)
        wb = load_workbook(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "Project"  # header
        keys = [r[1] for r in rows[1:]]
        assert "A-1" in keys
        assert "A-2" in keys

    def test_records_sorted_by_project_then_key(self, tmp_path):
        """Records are written sorted by project key, then issue key."""
        from openpyxl import load_workbook
        path = tmp_path / "sorted.xlsx"
        records = [
            _make_record("B-2", project="ARTB"),
            _make_record("A-1", project="ARTA"),
            _make_record("A-2", project="ARTA"),
        ]
        write_zero_day_excel(records, path)
        wb = load_workbook(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        keys = [r[1] for r in rows[1:]]
        assert keys == ["A-1", "A-2", "B-2"]

    def test_empty_records_writes_header_only(self, tmp_path):
        """Passing an empty list produces a file with only the header row."""
        from openpyxl import load_workbook
        path = tmp_path / "empty.xlsx"
        write_zero_day_excel([], path)
        wb = load_workbook(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1  # header only
        assert rows[0][0] == "Project"


def _make_report_record(
    key: str,
    first: "datetime | None" = None,
    closed: "datetime | None" = None,
    stage_minutes: "dict | None" = None,
) -> IssueRecord:
    """Create an IssueRecord for write_report_excel tests.

    Args:
        key:           Unique issue key.
        first:         First Date (start of active work), or None.
        closed:        Closed Date (completion), or None.
        stage_minutes: Stage-to-minutes mapping; defaults to empty dict.

    Returns:
        IssueRecord with project 'P' and the given date fields.
    """
    from datetime import datetime
    return IssueRecord(
        project="P", key=key, issuetype="Feature", status="Done",
        created=datetime(2025, 1, 1), component="",
        first_date=first, implementation_date=None,
        closed_date=closed,
        stage_minutes=stage_minutes or {},
        resolution="Done",
    )


class TestWriteReportExcel:
    """Tests for write_report_excel() — report XLSX with status group and cycle time columns."""

    _STAGES = ["Analysis", "In Dev", "Releasing"]

    def test_file_is_created(self, tmp_path):
        """An XLSX file is created at the given path."""
        from datetime import datetime
        path = tmp_path / "report.xlsx"
        write_report_excel([_make_report_record("A-1")], self._STAGES, path)
        assert path.exists()

    def test_creates_parent_dirs(self, tmp_path):
        """Intermediate parent directories are created if they do not exist."""
        path = tmp_path / "sub" / "report.xlsx"
        write_report_excel([], self._STAGES, path)
        assert path.exists()

    def test_header_contains_status_group(self, tmp_path):
        """The header row contains a 'Status Group' column."""
        from openpyxl import load_workbook
        path = tmp_path / "report.xlsx"
        write_report_excel([], self._STAGES, path)
        ws = load_workbook(path).active
        headers = [c.value for c in ws[1]]
        assert "Status Group" in headers

    def test_header_contains_cycle_time_columns(self, tmp_path):
        """The header row contains Cycle Time A and Cycle Time B columns."""
        from openpyxl import load_workbook
        path = tmp_path / "report.xlsx"
        write_report_excel([], self._STAGES, path)
        ws = load_workbook(path).active
        headers = [c.value for c in ws[1]]
        assert "Cycle Time (First->Closed)" in headers
        assert "Cycle Time B (days in Status)" in headers

    def test_header_contains_stage_columns(self, tmp_path):
        """Each stage name from the stages list appears as a column header."""
        from openpyxl import load_workbook
        path = tmp_path / "report.xlsx"
        write_report_excel([], self._STAGES, path)
        ws = load_workbook(path).active
        headers = [c.value for c in ws[1]]
        for stage in self._STAGES:
            assert stage in headers

    def test_status_group_done_for_closed_issue(self, tmp_path):
        """A closed issue receives the status group 'Done'."""
        from datetime import datetime
        from openpyxl import load_workbook
        path = tmp_path / "report.xlsx"
        rec = _make_report_record("A-1", datetime(2025, 1, 2), datetime(2025, 1, 10))
        write_report_excel([rec], self._STAGES, path)
        ws = load_workbook(path).active
        headers = [c.value for c in ws[1]]
        col_idx = headers.index("Status Group")
        row = list(ws.iter_rows(values_only=True))[1]
        assert row[col_idx] == "Done"

    def test_status_group_in_progress(self, tmp_path):
        """An issue with First Date but no Closed Date receives 'In Progress'."""
        from datetime import datetime
        from openpyxl import load_workbook
        path = tmp_path / "report.xlsx"
        rec = _make_report_record("A-2", first=datetime(2025, 1, 2), closed=None)
        write_report_excel([rec], self._STAGES, path)
        ws = load_workbook(path).active
        headers = [c.value for c in ws[1]]
        col_idx = headers.index("Status Group")
        row = list(ws.iter_rows(values_only=True))[1]
        assert row[col_idx] == "In Progress"

    def test_status_group_todo(self, tmp_path):
        """An issue without any transition dates receives 'To Do'."""
        from openpyxl import load_workbook
        path = tmp_path / "report.xlsx"
        rec = _make_report_record("A-3", first=None, closed=None)
        write_report_excel([rec], self._STAGES, path)
        ws = load_workbook(path).active
        headers = [c.value for c in ws[1]]
        col_idx = headers.index("Status Group")
        row = list(ws.iter_rows(values_only=True))[1]
        assert row[col_idx] == "To Do"

    def test_ct_a_correct_value(self, tmp_path):
        """Cycle Time A equals the calendar-day difference between First Date and Closed Date."""
        from datetime import datetime
        from openpyxl import load_workbook
        path = tmp_path / "report.xlsx"
        rec = _make_report_record("A-1", datetime(2025, 1, 1), datetime(2025, 1, 11))
        write_report_excel([rec], self._STAGES, path)
        ws = load_workbook(path).active
        headers = [c.value for c in ws[1]]
        col_idx = headers.index("Cycle Time (First->Closed)")
        row = list(ws.iter_rows(values_only=True))[1]
        assert row[col_idx] == 10.0

    def test_ct_b_correct_value(self, tmp_path):
        """Cycle Time B equals the sum of stage minutes (excl. last stage) divided by 1440."""
        from datetime import datetime
        from openpyxl import load_workbook
        path = tmp_path / "report.xlsx"
        # 1440 min in Analysis + 2880 min in In Dev; Releasing (last stage) excluded
        stage_minutes = {"Analysis": 1440, "In Dev": 2880, "Releasing": 999}
        rec = _make_report_record(
            "A-1", datetime(2025, 1, 1), datetime(2025, 1, 5), stage_minutes
        )
        write_report_excel([rec], self._STAGES, path)
        ws = load_workbook(path).active
        headers = [c.value for c in ws[1]]
        col_idx = headers.index("Cycle Time B (days in Status)")
        row = list(ws.iter_rows(values_only=True))[1]
        assert row[col_idx] == 3.0  # (1440 + 2880) / 1440

    def test_ct_empty_for_open_issue(self, tmp_path):
        """Cycle Time columns are empty for issues that lack First Date or Closed Date."""
        from openpyxl import load_workbook
        path = tmp_path / "report.xlsx"
        rec = _make_report_record("A-1", first=None, closed=None)
        write_report_excel([rec], self._STAGES, path)
        ws = load_workbook(path).active
        headers = [c.value for c in ws[1]]
        row = list(ws.iter_rows(values_only=True))[1]
        assert row[headers.index("Cycle Time (First->Closed)")] in (None, "")
        assert row[headers.index("Cycle Time B (days in Status)")] in (None, "")

    def test_empty_records_writes_header_only(self, tmp_path):
        """Passing an empty records list produces a file with only the header row."""
        from openpyxl import load_workbook
        path = tmp_path / "report.xlsx"
        write_report_excel([], self._STAGES, path)
        ws = load_workbook(path).active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1


class TestExportPng:
    """Tests for export_png() — PNG image export."""

    def test_calls_write_image_with_png_format(self, tmp_path, simple_figure):
        """export_png calls pio.write_image with format='png' and default dimensions."""
        out = tmp_path / "chart.png"
        with patch("build_reports.export.pio.write_image") as mock_write:
            export_png(simple_figure, out)
        mock_write.assert_called_once_with(
            simple_figure, str(out), format="png", width=1400, height=700
        )

    def test_creates_parent_dirs(self, tmp_path, simple_figure):
        """Intermediate parent directories are created if they do not exist."""
        out = tmp_path / "imgs" / "chart.png"
        with patch("build_reports.export.pio.write_image"):
            export_png(simple_figure, out)
        assert out.parent.exists()
