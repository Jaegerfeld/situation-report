# =============================================================================
# Autor:          Robert Seebauer
# Repository:     https://github.com/Jaegerfeld/situation-report
# KI-Unterstützung: Erstellt mit Unterstützung von Claude (Anthropic)
# Erstellt:       12.05.2026
# Geändert:       12.05.2026
# Lizenz:         BSD-3-Clause (siehe LICENSE)
#
# Fachliche Funktion:
#   Unit-Tests für build_reports.loader: Datums-Parser (_parse_dt, _parse_date),
#   Fehlerbehandlung in load_issue_times (fehlende Spalten, leere Zeilen),
#   load_cfd (ungültiges Datum), load_transitions (leere Zeilen) und
#   load_report_data mit Transitions-Pfad.
# =============================================================================

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest

from build_reports.loader import (
    _parse_date,
    _parse_dt,
    load_cfd,
    load_issue_times,
    load_report_data,
    load_transitions,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_ISSUE_TIMES_HEADER = [
    "Project", "Group", "Key", "Issuetype", "Status", "Created Date",
    "Component", "Category", "First Date", "Implementation Date", "Closed Date",
    "Analysis", "Dev",
    "Resolution",
]

_CFD_HEADER = ["Day", "Analysis", "Dev"]


def _make_issue_times(path: Path, data_rows: list[list] | None = None) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_ISSUE_TIMES_HEADER)
    for row in (data_rows or []):
        ws.append(row)
    wb.save(path)


def _make_cfd(path: Path, data_rows: list[list] | None = None) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_CFD_HEADER)
    for row in (data_rows or []):
        ws.append(row)
    wb.save(path)


def _make_transitions(path: Path, data_rows: list[list] | None = None) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Key", "Transition", "Timestamp"])
    for row in (data_rows or []):
        ws.append(row)
    wb.save(path)


# ---------------------------------------------------------------------------
# _parse_dt
# ---------------------------------------------------------------------------

class TestParseDt:
    def test_none_returns_none(self):
        assert _parse_dt(None) is None

    def test_empty_string_returns_none(self):
        assert _parse_dt("") is None

    def test_datetime_object_returned_as_is(self):
        dt = datetime(2025, 6, 15, 10, 30, 0)
        assert _parse_dt(dt) is dt

    def test_valid_string_parsed(self):
        result = _parse_dt("15.06.2025 10:30:00")
        assert result == datetime(2025, 6, 15, 10, 30, 0)

    def test_invalid_string_returns_none(self):
        assert _parse_dt("not-a-date") is None

    def test_non_string_non_datetime_returns_none(self):
        assert _parse_dt(42) is None


# ---------------------------------------------------------------------------
# _parse_date
# ---------------------------------------------------------------------------

class TestParseDate:
    def test_none_returns_none(self):
        assert _parse_date(None) is None

    def test_empty_string_returns_none(self):
        assert _parse_date("") is None

    def test_datetime_returns_date(self):
        dt = datetime(2025, 6, 15, 10, 30)
        assert _parse_date(dt) == date(2025, 6, 15)

    def test_date_object_returned_as_is(self):
        d = date(2025, 6, 15)
        assert _parse_date(d) is d

    def test_valid_string_parsed(self):
        assert _parse_date("15.06.2025") == date(2025, 6, 15)

    def test_invalid_string_returns_none(self):
        assert _parse_date("not-a-date") is None

    def test_non_string_non_date_returns_none(self):
        assert _parse_date(99) is None


# ---------------------------------------------------------------------------
# load_issue_times
# ---------------------------------------------------------------------------

class TestLoadIssueTimes:
    def test_missing_closed_date_column_raises(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Project", "Key", "Resolution"])
        path = tmp_path / "IT.xlsx"
        wb.save(path)
        with pytest.raises(ValueError, match="Unexpected IssueTimes format"):
            load_issue_times(path)

    def test_empty_row_skipped(self, tmp_path):
        path = tmp_path / "IT.xlsx"
        valid_row = ["PROJ", None, "T-1", "Story", "Done",
                     "01.01.2025 09:00:00", "", "", None, None, None, 0, 0, "Done"]
        _make_issue_times(path, data_rows=[valid_row, [None] * len(_ISSUE_TIMES_HEADER)])
        records, stages = load_issue_times(path)
        assert len(records) == 1

    def test_stages_inferred_from_header(self, tmp_path):
        path = tmp_path / "IT.xlsx"
        _make_issue_times(path)
        _, stages = load_issue_times(path)
        assert stages == ["Analysis", "Dev"]


# ---------------------------------------------------------------------------
# load_cfd
# ---------------------------------------------------------------------------

class TestLoadCfd:
    def test_none_day_row_skipped(self, tmp_path):
        path = tmp_path / "CFD.xlsx"
        _make_cfd(path, data_rows=[
            ["not-a-date", 5, 3],
            ["15.06.2025", 5, 3],
        ])
        records, _ = load_cfd(path)
        assert len(records) == 1

    def test_empty_row_skipped(self, tmp_path):
        path = tmp_path / "CFD.xlsx"
        _make_cfd(path, data_rows=[
            ["15.06.2025", 5, 3],
            [None, None, None],
        ])
        records, _ = load_cfd(path)
        assert len(records) == 1


# ---------------------------------------------------------------------------
# load_transitions
# ---------------------------------------------------------------------------

class TestLoadTransitions:
    def test_empty_row_skipped(self, tmp_path):
        path = tmp_path / "Transitions.xlsx"
        _make_transitions(path, data_rows=[
            ["T-1", "Analysis", "15.06.2025 10:00:00"],
            [None, None, None],
        ])
        entries = load_transitions(path)
        assert len(entries) == 1


# ---------------------------------------------------------------------------
# load_report_data
# ---------------------------------------------------------------------------

class TestLoadReportData:
    def test_with_transitions_path(self, tmp_path):
        it = tmp_path / "ART_A_IssueTimes.xlsx"
        _make_issue_times(it)
        tr = tmp_path / "ART_A_Transitions.xlsx"
        _make_transitions(tr, data_rows=[["T-1", "Analysis", "15.06.2025 10:00:00"]])
        data = load_report_data(it, transitions_path=tr)
        assert len(data.transitions) == 1
        assert data.transitions[0].key == "T-1"


# ---------------------------------------------------------------------------
# metrics registry — duplicate registration
# ---------------------------------------------------------------------------

class TestMetricsRegistry:
    def test_duplicate_registration_raises(self):
        from build_reports.metrics import _REGISTRY, register
        from build_reports.metrics.base import MetricPlugin, MetricResult

        class _TmpPlugin(MetricPlugin):
            metric_id = "_test_dup_tmp"

            def compute(self, data, terminology) -> MetricResult:
                return MetricResult(metric_id=self.metric_id, stats={}, chart_data={})

            def render(self, result, terminology) -> list:
                return []

        plugin = _TmpPlugin()
        try:
            register(plugin)
            with pytest.raises(ValueError, match="already registered"):
                register(plugin)
        finally:
            _REGISTRY.pop("_test_dup_tmp", None)
