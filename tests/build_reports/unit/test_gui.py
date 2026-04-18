# =============================================================================
# Autor:          Robert Seebauer
# Repository:     https://github.com/Jaegerfeld/situation-report
# KI-Unterstützung: Erstellt mit Unterstützung von Claude (Anthropic)
# Erstellt:       16.04.2026
# Geändert:       16.04.2026
# Lizenz:         BSD-3-Clause (siehe LICENSE)
#
# Fachliche Funktion:
#   Unit-Tests für die Display-unabhängigen Hilfsfunktionen in gui.py:
#   _parse_date_safe, _split_csv, _build_combined_html und _TRANSLATIONS.
#   Der tkinter-Teil (BuildReportsApp) wird hier nicht instanziiert, da dies
#   eine laufende Anzeige erfordern würde.
# =============================================================================

from __future__ import annotations

from datetime import date

import plotly.graph_objects as go
import pytest

from build_reports.gui import (
    LANG_DE, LANG_EN, _T,
    _build_combined_html, _build_template_dict, _check_stage_consistency,
    _default_date_range, _parse_date_safe, _parse_template_dict,
    _read_available_filters, _split_csv, _TEMPLATE_VERSION,
)


class TestParseDateSafe:
    """Tests for _parse_date_safe() — lenient date string parsing for GUI fields."""

    def test_empty_string_returns_none(self):
        """An empty string returns None (field left blank by the user)."""
        assert _parse_date_safe("") is None

    def test_whitespace_only_returns_none(self):
        """A whitespace-only string is treated as blank and returns None."""
        assert _parse_date_safe("   ") is None

    def test_valid_date_parsed(self):
        """A valid ISO date string is parsed to a date object."""
        assert _parse_date_safe("2025-06-15") == date(2025, 6, 15)

    def test_valid_date_with_surrounding_whitespace(self):
        """Leading and trailing whitespace is stripped before parsing."""
        assert _parse_date_safe("  2025-01-01  ") == date(2025, 1, 1)

    def test_invalid_format_raises_value_error(self):
        """A non-ISO date format raises ValueError."""
        with pytest.raises(ValueError):
            _parse_date_safe("15.06.2025")

    def test_invalid_date_raises_value_error(self):
        """An impossible date (month 13) raises ValueError."""
        with pytest.raises(ValueError):
            _parse_date_safe("2025-13-01")


class TestSplitCsv:
    """Tests for _split_csv() — comma-separated entry field parsing."""

    def test_empty_string_returns_empty_list(self):
        """An empty string produces an empty list."""
        assert _split_csv("") == []

    def test_whitespace_only_returns_empty_list(self):
        """A whitespace-only string produces an empty list."""
        assert _split_csv("   ") == []

    def test_single_item(self):
        """A single item without commas is returned in a one-element list."""
        assert _split_csv("ARTA") == ["ARTA"]

    def test_multiple_items(self):
        """Multiple comma-separated items are split and returned as a list."""
        assert _split_csv("ARTA, ARTB, ARTC") == ["ARTA", "ARTB", "ARTC"]

    def test_strips_whitespace_from_items(self):
        """Whitespace around individual items is stripped."""
        assert _split_csv("  Feature ,  Bug  ") == ["Feature", "Bug"]

    def test_ignores_empty_segments(self):
        """Empty segments (consecutive commas) are ignored."""
        assert _split_csv("ARTA,,ARTB") == ["ARTA", "ARTB"]

    def test_trailing_comma_ignored(self):
        """A trailing comma does not produce an empty item."""
        assert _split_csv("ARTA,") == ["ARTA"]


class TestBuildCombinedHtml:
    """Tests for _build_combined_html() — multi-figure HTML document generation."""

    def test_returns_string(self):
        """Output is a string."""
        fig = go.Figure(go.Scatter(x=[1, 2], y=[1, 2]))
        html = _build_combined_html([fig])
        assert isinstance(html, str)

    def test_html_contains_doctype(self):
        """The output starts with a valid HTML doctype declaration."""
        fig = go.Figure()
        html = _build_combined_html([fig])
        assert "<!DOCTYPE html>" in html

    def test_first_figure_includes_plotlyjs(self):
        """The first figure embeds the Plotly.js CDN reference."""
        fig = go.Figure()
        html = _build_combined_html([fig])
        # cdn reference appears when include_plotlyjs='cdn'
        assert "cdn.plot.ly" in html or "plotly" in html.lower()

    def test_multiple_figures_all_embedded(self):
        """All figures are embedded as separate plotly divs."""
        figs = [go.Figure(go.Scatter(x=[i], y=[i])) for i in range(3)]
        html = _build_combined_html(figs)
        # Each figure generates a unique div id; count plotly-graph-div occurrences
        assert html.count("plotly-graph-div") >= 3

    def test_empty_list_returns_valid_html(self):
        """Passing an empty figures list produces a valid (empty) HTML document."""
        html = _build_combined_html([])
        assert "<html>" in html
        assert "<body>" in html

    def test_section_break_heading_inserted(self):
        """A section_breaks entry inserts an h2 heading before the corresponding figure."""
        figs = [go.Figure(), go.Figure()]
        html = _build_combined_html(figs, section_breaks={0: "Flow Time"})
        assert "Flow Time" in html
        assert "metric-heading" in html

    def test_section_break_only_at_correct_index(self):
        """The section heading appears after earlier figure divs, not before."""
        figs = [go.Figure(), go.Figure()]
        html = _build_combined_html(figs, section_breaks={1: "Throughput"})
        # Heading must appear before the second figure block
        idx_heading = html.index("Throughput")
        # At least one plotly div precedes the heading
        idx_first_div = html.index("plotly-graph-div")
        assert idx_first_div < idx_heading

    def test_no_section_breaks_produces_no_h2(self):
        """When section_breaks is None, no h2 headings appear in the output."""
        fig = go.Figure()
        html = _build_combined_html([fig], section_breaks=None)
        assert "<h2" not in html


def _make_issuetimes(tmp_path, stages: list[str], name="it.xlsx"):
    """Create a minimal IssueTimes XLSX file with the given stage columns.

    Args:
        tmp_path: pytest temporary directory.
        stages:   Stage column names to include between Closed Date and Resolution.
        name:     Filename within tmp_path (default 'it.xlsx').

    Returns:
        Path to the created XLSX file.
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Project", "Key", "Issuetype", "Status", "Created Date",
               "Component", "Category", "First Date", "Implementation Date",
               "Closed Date"] + stages + ["Resolution"])
    ws.append(["ART", "ART-1", "Feature", "Done", None, "", "", None, None, None]
              + [0] * len(stages) + ["Done"])
    path = tmp_path / name
    wb.save(path)
    return path


def _make_cfd(tmp_path, stages: list[str], name="cfd.xlsx"):
    """Create a minimal CFD XLSX file with the given stage columns.

    Args:
        tmp_path: pytest temporary directory.
        stages:   Stage column names (appended after the 'Day' column).
        name:     Filename within tmp_path (default 'cfd.xlsx').

    Returns:
        Path to the created XLSX file.
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Day"] + stages)
    ws.append(["2025-01-01"] + [0] * len(stages))
    path = tmp_path / name
    wb.save(path)
    return path


class TestCheckStageConsistency:
    """Tests for _check_stage_consistency() — stage column comparison between files."""

    def test_consistent_returns_empty_lists(self, tmp_path):
        """Identical stage columns in both files return two empty lists."""
        stages = ["Analysis", "Implementation", "Done"]
        it = _make_issuetimes(tmp_path, stages)
        cfd = _make_cfd(tmp_path, stages)
        only_it, only_cfd = _check_stage_consistency(it, cfd)
        assert only_it == []
        assert only_cfd == []

    def test_stage_missing_in_cfd(self, tmp_path):
        """A stage present in IssueTimes but absent in CFD appears in only_in_it."""
        it = _make_issuetimes(tmp_path, ["Analysis", "Implementation", "Done"])
        cfd = _make_cfd(tmp_path, ["Analysis", "Done"])
        only_it, only_cfd = _check_stage_consistency(it, cfd)
        assert only_it == ["Implementation"]
        assert only_cfd == []

    def test_extra_stage_in_cfd(self, tmp_path):
        """A stage present in CFD but absent in IssueTimes appears in only_in_cfd."""
        it = _make_issuetimes(tmp_path, ["Analysis", "Done"])
        cfd = _make_cfd(tmp_path, ["Analysis", "Implementation", "Done"])
        only_it, only_cfd = _check_stage_consistency(it, cfd)
        assert only_it == []
        assert only_cfd == ["Implementation"]

    def test_both_mismatches(self, tmp_path):
        """Mismatches in both directions are reported simultaneously."""
        it = _make_issuetimes(tmp_path, ["A", "B"])
        cfd = _make_cfd(tmp_path, ["A", "C"])
        only_it, only_cfd = _check_stage_consistency(it, cfd)
        assert only_it == ["B"]
        assert only_cfd == ["C"]

    def test_translation_keys_present(self):
        """Translation keys for stage consistency log messages exist in both languages."""
        for lang in (LANG_DE, LANG_EN):
            assert "log_check_ok" in _T[lang]
            assert "log_check_miss_cfd" in _T[lang]
            assert "log_check_miss_it" in _T[lang]


class TestReadAvailableFilters:
    """Tests for _read_available_filters() — project and issuetype extraction from XLSX."""

    def test_returns_sorted_projects(self, tmp_path):
        """Projects extracted from the XLSX are sorted and deduplicated."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Project", "Key", "Issuetype", "Status"])
        ws.append(["ARTB", "ARTB-1", "Feature", "Done"])
        ws.append(["ARTA", "ARTA-1", "Bug", "Open"])
        ws.append(["ARTA", "ARTA-2", "Feature", "Done"])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        projects, _ = _read_available_filters(path)
        assert projects == ["ARTA", "ARTB"]

    def test_returns_sorted_issuetypes(self, tmp_path):
        """Issue types extracted from the XLSX are sorted and deduplicated."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Project", "Key", "Issuetype", "Status"])
        ws.append(["ART", "ART-1", "Story", "Done"])
        ws.append(["ART", "ART-2", "Feature", "Done"])
        ws.append(["ART", "ART-3", "Story", "Open"])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        _, issuetypes = _read_available_filters(path)
        assert issuetypes == ["Feature", "Story"]

    def test_empty_file_returns_empty_lists(self, tmp_path):
        """A file with only the header row produces empty project and issuetype lists."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Project", "Key", "Issuetype", "Status"])
        path = tmp_path / "empty.xlsx"
        wb.save(path)
        projects, issuetypes = _read_available_filters(path)
        assert projects == []
        assert issuetypes == []


class TestDefaultDateRange:
    """Tests for _default_date_range() — default GUI filter date range computation."""

    def test_returns_tuple_of_two_dates(self):
        """The function returns a tuple of two date objects."""
        from_d, to_d = _default_date_range()
        assert isinstance(from_d, date)
        assert isinstance(to_d, date)

    def test_to_is_today(self):
        """to_date equals today's date."""
        _, to_d = _default_date_range()
        assert to_d == date.today()

    def test_from_is_365_days_before_today(self):
        """from_date is exactly 365 days before today."""
        from datetime import timedelta
        from_d, _ = _default_date_range()
        assert from_d == date.today() - timedelta(days=365)

    def test_from_before_to(self):
        """from_date is strictly before to_date."""
        from_d, to_d = _default_date_range()
        assert from_d < to_d


class TestTranslations:
    """Tests for the _T translation table — key completeness and consistency."""

    def test_both_languages_present(self):
        """Both DE and EN language codes are registered in _T."""
        assert LANG_DE in _T
        assert LANG_EN in _T

    def test_same_keys_in_both_languages(self):
        """Both languages have the exact same set of translation keys."""
        assert set(_T[LANG_DE].keys()) == set(_T[LANG_EN].keys())

    def test_german_window_title(self):
        """The German window title is 'build_reports'."""
        assert _T[LANG_DE]["window_title"] == "build_reports"

    def test_german_options_menu_label(self):
        """The German options menu label is 'Optionen'."""
        assert _T[LANG_DE]["menu_options"] == "Optionen"

    def test_english_options_menu_label(self):
        """The English options menu label is 'Options'."""
        assert _T[LANG_EN]["menu_options"] == "Options"

    def test_no_empty_values(self):
        """No translation key maps to an empty string in any language."""
        for lang, entries in _T.items():
            for key, value in entries.items():
                assert value, f"Empty translation: [{lang}][{key}]"

    def test_date_picker_keys_present(self):
        """Translation keys required for the calendar picker dialog exist."""
        for lang in (LANG_DE, LANG_EN):
            assert "dlg_pick_date" in _T[lang]
            assert "btn_cal" in _T[lang]
            assert "btn_ok" in _T[lang]
            assert "btn_last_365" in _T[lang]

    def test_filter_picker_keys_present(self):
        """Translation keys required for the filter picker dialogs exist."""
        for lang in (LANG_DE, LANG_EN):
            assert "dlg_projects" in _T[lang]
            assert "dlg_issuetypes" in _T[lang]
            assert "btn_pick" in _T[lang]

    def test_template_keys_present(self):
        """Translation keys required for template save/load menus exist."""
        for lang in (LANG_DE, LANG_EN):
            assert "menu_template" in _T[lang]
            assert "menu_tpl_save" in _T[lang]
            assert "menu_tpl_load" in _T[lang]
            assert "log_tpl_saved" in _T[lang]
            assert "log_tpl_loaded" in _T[lang]
            assert "log_tpl_error" in _T[lang]

    def test_tooltip_keys_present(self):
        """All required tooltip translation keys exist in both languages."""
        required = [
            "tip_issue_times", "tip_cfd", "tip_pi_config", "tip_browse",
            "tip_from", "tip_to", "tip_cal", "tip_last_365",
            "tip_projects", "tip_issuetypes", "tip_pick",
            "tip_ct_a", "tip_ct_b", "tip_show", "tip_pdf",
        ]
        for lang in (LANG_DE, LANG_EN):
            for key in required:
                assert key in _T[lang], f"Missing tooltip key [{lang}][{key}]"

    def test_metric_tooltip_keys_present(self):
        """If a metric tooltip key exists in DE, the same key must exist in EN."""
        from build_reports.metrics import all_metrics
        for plugin in all_metrics():
            tip_key = f"tip_metric_{plugin.metric_id}"
            # Only check keys that are defined (not all metrics need one)
            if tip_key in _T[LANG_DE]:
                assert tip_key in _T[LANG_EN], (
                    f"Tooltip key {tip_key} present in DE but missing in EN"
                )


# ---------------------------------------------------------------------------
# Helper for template tests
# ---------------------------------------------------------------------------

def _sample_template(**overrides) -> dict:
    """Return a minimal valid template dict, with optional field overrides.

    Args:
        **overrides: Key-value pairs to override in the base template.

    Returns:
        Dict with all fields required by _build_template_dict.
    """
    base = dict(
        issue_times="/data/it.xlsx",
        cfd="/data/cfd.xlsx",
        pi_config="",
        from_date="2024-01-01",
        to_date="2024-12-31",
        projects="ARTA, ARTB",
        issuetypes="Feature",
        terminology="safe",
        ct_method="A",
        metrics={"flow_time": True, "throughput": False},
        language="de",
    )
    base.update(overrides)
    return base


class TestBuildTemplateDict:
    """Tests for _build_template_dict() — GUI state serialisation."""

    def test_contains_version(self):
        """The template dict always contains an integer 'version' key."""
        tpl = _build_template_dict(**_sample_template())
        assert "version" in tpl
        assert isinstance(tpl["version"], int)

    def test_roundtrip_fields(self):
        """All string fields survive a build → parse roundtrip unchanged."""
        state = _sample_template()
        tpl = _build_template_dict(**state)
        for key in ("issue_times", "cfd", "pi_config", "from_date", "to_date",
                    "projects", "issuetypes", "terminology", "ct_method", "language"):
            assert tpl[key] == state[key]

    def test_metrics_preserved(self):
        """The metrics dict is stored exactly as provided."""
        metrics = {"flow_time": True, "throughput": False}
        tpl = _build_template_dict(**_sample_template(metrics=metrics))
        assert tpl["metrics"] == metrics

    def test_empty_paths_allowed(self):
        """Empty strings for file paths are allowed and stored as-is."""
        tpl = _build_template_dict(**_sample_template(issue_times="", cfd=""))
        assert tpl["issue_times"] == ""
        assert tpl["cfd"] == ""


class TestParseTemplateDict:
    """Tests for _parse_template_dict() — template JSON loading and validation."""

    def test_valid_roundtrip(self):
        """A built template can be parsed back to the original field values."""
        state = _sample_template()
        tpl = _build_template_dict(**state)
        parsed = _parse_template_dict(tpl)
        for key in ("issue_times", "cfd", "pi_config", "from_date", "to_date",
                    "projects", "issuetypes", "terminology", "ct_method", "language"):
            assert parsed[key] == state[key]

    def test_missing_keys_get_defaults(self):
        """Missing keys fall back to safe defaults so older templates remain loadable."""
        parsed = _parse_template_dict({"version": 1})
        assert parsed["issue_times"] == ""
        assert parsed["cfd"] == ""
        assert parsed["pi_config"] == ""
        assert parsed["from_date"] == ""
        assert parsed["to_date"] == ""
        assert isinstance(parsed["metrics"], dict)

    def test_non_dict_raises_value_error(self):
        """Passing a non-dict (e.g. a list) raises ValueError."""
        with pytest.raises(ValueError):
            _parse_template_dict([1, 2, 3])

    def test_future_version_raises_value_error(self):
        """A template version higher than supported raises ValueError."""
        with pytest.raises(ValueError, match="version"):
            _parse_template_dict({"version": 9999})

    def test_metrics_is_dict(self):
        """The metrics field is parsed as a dict."""
        parsed = _parse_template_dict({"version": 1, "metrics": {"flow_time": True}})
        assert parsed["metrics"] == {"flow_time": True}

    def test_json_serialisable(self):
        """The built template dict is JSON-serialisable without errors."""
        import json
        state = _sample_template()
        tpl = _build_template_dict(**state)
        # Must not raise
        json.dumps(tpl)
