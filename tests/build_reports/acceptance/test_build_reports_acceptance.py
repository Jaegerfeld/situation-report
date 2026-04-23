# =============================================================================
# Autor:          Robert Seebauer
# Repository:     https://github.com/Jaegerfeld/situation-report
# KI-Unterstützung: Erstellt mit Unterstützung von Claude (Anthropic)
# Erstellt:       15.04.2026
# Geändert:       18.04.2026
# Lizenz:         BSD-3-Clause (siehe LICENSE)
#
# Fachliche Funktion:
#   Acceptance-Tests für build_reports gegen den realen ART_A-Datensatz.
#   Prüft, dass loader, filters und alle Metrik-Plugins mit echten Daten
#   korrekt funktionieren: sinnvolle Statistikwerte, vollständige Datenladung,
#   renderfähige Figures.
# =============================================================================

from datetime import date
from pathlib import Path

import pytest

from build_reports.cli import run_reports
from build_reports.export import export_pdf, export_png
from build_reports.filters import FilterConfig, apply_filters
from build_reports.loader import load_report_data
from build_reports.metrics.cfd import CfdMetric
from build_reports.metrics.flow_distribution import FlowDistributionMetric
from build_reports.metrics.flow_load import FlowLoadMetric
from build_reports.metrics.flow_time import FlowTimeMetric
from build_reports.metrics.flow_velocity import FlowVelocityMetric
from build_reports.terminology import SAFE

TESTDATA_DIR = Path(__file__).parent.parent.parent / "testdata" / "ART_A"
ISSUE_TIMES = TESTDATA_DIR / "ART_A_IssueTimes.xlsx"
CFD = TESTDATA_DIR / "ART_A_CFD.xlsx"


@pytest.fixture(scope="module")
def art_a_data():
    """Load the real ART_A dataset once for all acceptance tests."""
    return load_report_data(ISSUE_TIMES, CFD)


class TestLoader:
    """Acceptance tests for load_report_data() against the real ART_A dataset."""

    def test_loads_issues(self, art_a_data):
        """At least one issue is loaded from the IssueTimes file."""
        assert len(art_a_data.issues) > 0

    def test_loads_cfd(self, art_a_data):
        """At least one CFD record is loaded from the CFD file."""
        assert len(art_a_data.cfd) > 0

    def test_stages_not_empty(self, art_a_data):
        """The workflow stage list derived from the dataset is non-empty."""
        assert len(art_a_data.stages) > 0

    def test_source_prefix(self, art_a_data):
        """source_prefix is derived from the filename and equals 'ART_A'."""
        assert art_a_data.source_prefix == "ART_A"

    def test_issues_have_keys(self, art_a_data):
        """Every loaded issue has a non-empty key string."""
        assert all(i.key for i in art_a_data.issues)

    def test_issues_have_project(self, art_a_data):
        """Every loaded issue has a non-empty project string."""
        assert all(i.project for i in art_a_data.issues)

    def test_stage_minutes_keys_match_stages(self, art_a_data):
        """stage_minutes dict keys on every issue match the dataset's stage list exactly."""
        for issue in art_a_data.issues:
            assert set(issue.stage_minutes.keys()) == set(art_a_data.stages)


class TestFiltersOnRealData:
    """Acceptance tests for apply_filters() — date range filtering on the real ART_A dataset."""

    def test_date_filter_reduces_issues(self, art_a_data):
        """A from_date filter returns fewer issues than the unfiltered dataset."""
        cfg = FilterConfig(from_date=date(2026, 1, 1))
        filtered = apply_filters(art_a_data, cfg)
        assert len(filtered.issues) < len(art_a_data.issues)

    def test_date_filter_closed_issues_in_range(self, art_a_data):
        """Issues with a Closed Date are within the filter range; open issues pass through."""
        cfg = FilterConfig(from_date=date(2026, 1, 1), to_date=date(2026, 12, 31))
        filtered = apply_filters(art_a_data, cfg)
        for issue in filtered.issues:
            if issue.closed_date is not None:
                assert issue.closed_date.date() >= date(2026, 1, 1)
                assert issue.closed_date.date() <= date(2026, 12, 31)

    def test_cfd_filter_reduces_records(self, art_a_data):
        """A from_date filter returns fewer CFD records than the unfiltered dataset."""
        cfg = FilterConfig(from_date=date(2026, 1, 1))
        filtered = apply_filters(art_a_data, cfg)
        assert len(filtered.cfd) < len(art_a_data.cfd)


class TestFlowTimeOnRealData:
    """Acceptance tests for FlowTimeMetric against the real ART_A dataset."""

    def test_compute_returns_count(self, art_a_data):
        """compute() produces a positive issue count from the real dataset."""
        metric = FlowTimeMetric()
        result = metric.compute(art_a_data, SAFE)
        assert result.stats.get("count", 0) > 0

    def test_median_is_positive(self, art_a_data):
        """Median flow time is greater than zero for real issues."""
        metric = FlowTimeMetric()
        result = metric.compute(art_a_data, SAFE)
        assert result.stats["median"] > 0

    def test_min_less_than_max(self, art_a_data):
        """Minimum flow time is not greater than the maximum flow time."""
        metric = FlowTimeMetric()
        result = metric.compute(art_a_data, SAFE)
        assert result.stats["min"] <= result.stats["max"]

    def test_render_produces_two_figures(self, art_a_data):
        """render() returns exactly two figures for the real dataset."""
        metric = FlowTimeMetric()
        result = metric.compute(art_a_data, SAFE)
        figures = metric.render(result, SAFE)
        assert len(figures) == 2


class TestFlowVelocityOnRealData:
    """Acceptance tests for FlowVelocityMetric against the real ART_A dataset."""

    def test_count_positive(self, art_a_data):
        """Total closed issue count is positive for the real dataset."""
        result = FlowVelocityMetric().compute(art_a_data, SAFE)
        assert result.stats.get("count", 0) > 0

    def test_render_three_figures(self, art_a_data):
        """render() returns exactly three figures: daily hist, weekly line, PI bar."""
        metric = FlowVelocityMetric()
        result = metric.compute(art_a_data, SAFE)
        assert len(metric.render(result, SAFE)) == 3

    def test_avg_per_week_positive(self, art_a_data):
        """Average items closed per week is positive for the real dataset."""
        result = FlowVelocityMetric().compute(art_a_data, SAFE)
        assert result.stats["avg_per_week"] > 0


class TestFlowLoadOnRealData:
    """Acceptance tests for FlowLoadMetric against the real ART_A dataset."""

    def test_open_count_positive(self, art_a_data):
        """Number of currently open (in-progress) issues is positive."""
        result = FlowLoadMetric().compute(art_a_data, SAFE)
        assert result.stats.get("open_count", 0) > 0

    def test_render_one_figure(self, art_a_data):
        """render() returns exactly one figure for the real dataset."""
        metric = FlowLoadMetric()
        result = metric.compute(art_a_data, SAFE)
        figs = metric.render(result, SAFE)
        assert len(figs) == 1


class TestCfdOnRealData:
    """Acceptance tests for CfdMetric against the real ART_A dataset."""

    def test_days_positive(self, art_a_data):
        """Number of days covered by the CFD is positive for the real dataset."""
        result = CfdMetric().compute(art_a_data, SAFE)
        assert result.stats.get("days", 0) > 0

    def test_render_one_figure(self, art_a_data):
        """render() returns exactly one CFD figure for the real dataset."""
        metric = CfdMetric()
        result = metric.compute(art_a_data, SAFE)
        figs = metric.render(result, SAFE)
        assert len(figs) == 1

    def test_ratio_positive(self, art_a_data):
        """Throughput-to-arrival ratio is positive for the real dataset."""
        result = CfdMetric().compute(art_a_data, SAFE)
        assert result.stats.get("ratio", 0) > 0


class TestFlowDistributionOnRealData:
    """Acceptance tests for FlowDistributionMetric against the real ART_A dataset."""

    def test_total_matches_issue_count(self, art_a_data):
        """Distribution total equals the total number of loaded issues."""
        result = FlowDistributionMetric().compute(art_a_data, SAFE)
        assert result.stats["total"] == len(art_a_data.issues)

    def test_render_one_figure_with_two_pies_and_one_bar(self, art_a_data):
        """render() returns one figure with two pie traces and one bar trace."""
        metric = FlowDistributionMetric()
        result = metric.compute(art_a_data, SAFE)
        figs = metric.render(result, SAFE)
        assert len(figs) == 1
        assert len([t for t in figs[0].data if t.type == "pie"]) == 2
        assert len([t for t in figs[0].data if t.type == "bar"]) == 1


class TestCliPipelineOnRealData:
    """Acceptance tests for the run_reports() CLI pipeline against the real ART_A dataset."""

    def test_run_reports_all_metrics_produces_pdf(self, tmp_path):
        """Full pipeline via run_reports() against real ART_A data exports a non-empty PDF."""
        out = tmp_path / "art_a_report.pdf"
        logged: list[str] = []
        run_reports(
            issue_times=ISSUE_TIMES,
            cfd=CFD,
            output_pdf=out,
            log=logged.append,
        )
        assert out.exists(), "PDF not created"
        assert out.stat().st_size > 0, "PDF is empty"
        assert any("Saved" in m for m in logged)

    def test_run_reports_single_metric(self, tmp_path):
        """run_reports with a single metric produces fewer figures than all metrics."""
        out_single = tmp_path / "single.pdf"
        out_all = tmp_path / "all.pdf"
        counts: list[int] = []

        def counting_log(msg: str) -> None:
            """Capture figure counts from log messages of the form '  N figure(s) ...'.

            Args:
                msg: Log message emitted by run_reports during pipeline execution.
            """
            if "figure(s)" in msg:
                counts.append(int(msg.strip().split()[1]))

        run_reports(
            issue_times=ISSUE_TIMES,
            cfd=CFD,
            metrics=["flow_time"],
            output_pdf=out_single,
            log=counting_log,
        )
        run_reports(
            issue_times=ISSUE_TIMES,
            cfd=CFD,
            output_pdf=out_all,
            log=counting_log,
        )
        assert counts[0] < counts[-1], "Single metric should produce fewer figures than all"

    def test_run_reports_date_filter_accepted(self, tmp_path):
        """run_reports accepts date filters without crashing."""
        out = tmp_path / "filtered.pdf"
        run_reports(
            issue_times=ISSUE_TIMES,
            cfd=CFD,
            from_date=date(2026, 1, 1),
            output_pdf=out,
            log=lambda *_: None,
        )
        assert out.exists()

    def test_run_reports_creates_xlsx_alongside_pdf(self, tmp_path):
        """run_reports creates a report XLSX with the same stem as the PDF."""
        out = tmp_path / "art_a.pdf"
        run_reports(
            issue_times=ISSUE_TIMES,
            cfd=CFD,
            output_pdf=out,
            log=lambda *_: None,
        )
        xlsx = out.with_suffix(".xlsx")
        assert xlsx.exists(), "XLSX not created alongside PDF"
        assert xlsx.stat().st_size > 0, "XLSX is empty"

    def test_report_xlsx_contains_status_group_and_cycle_times(self, tmp_path):
        """The report XLSX contains Status Group, Cycle Time A and Cycle Time B columns."""
        from openpyxl import load_workbook
        out = tmp_path / "art_a.pdf"
        run_reports(
            issue_times=ISSUE_TIMES,
            cfd=CFD,
            output_pdf=out,
            log=lambda *_: None,
        )
        ws = load_workbook(out.with_suffix(".xlsx")).active
        headers = [c.value for c in ws[1]]
        assert "Status Group" in headers
        assert "Cycle Time (First->Closed)" in headers
        assert "Cycle Time B (days in Status)" in headers


class TestGuiOnRealData:
    """Acceptance tests for BuildReportsApp GUI class against real file paths."""

    def test_gui_instantiates_with_all_metrics(self):
        """BuildReportsApp can be instantiated and all metric plugins are registered."""
        from build_reports.gui import BuildReportsApp
        app = BuildReportsApp()
        try:
            assert len(app._metric_vars) > 0, "No metric checkboxes registered"
            from build_reports.metrics import all_metrics
            assert len(app._metric_vars) == len(all_metrics())
        finally:
            app.destroy()

    def test_read_inputs_fails_without_file(self):
        """_read_inputs returns None when no IssueTimes file is set."""
        from build_reports.gui import BuildReportsApp
        app = BuildReportsApp()
        try:
            result = app._read_inputs()
            assert result is None
        finally:
            app.destroy()

    def test_read_inputs_returns_dict_with_valid_file(self):
        """_read_inputs returns a complete dict when a valid file path is set."""
        from build_reports.gui import BuildReportsApp
        app = BuildReportsApp()
        try:
            app._issue_times_var.set(str(ISSUE_TIMES))
            app._cfd_var.set(str(CFD))
            result = app._read_inputs()
            assert result is not None
            assert result["issue_times"] == ISSUE_TIMES
            assert result["cfd"] == CFD
            # all metrics selected → full list returned (equivalent to None for run_reports)
            from build_reports.metrics import all_metrics
            assert set(result["metrics"]) == {p.metric_id for p in all_metrics()}
        finally:
            app.destroy()

    def test_deselect_all_metrics_sets_specific_list(self):
        """After deselecting all and re-selecting flow_time, only flow_time is returned."""
        from build_reports.gui import BuildReportsApp
        app = BuildReportsApp()
        try:
            app._issue_times_var.set(str(ISSUE_TIMES))
            app._deselect_all_metrics()
            app._metric_vars["flow_time"].set(True)
            result = app._read_inputs()
            assert result is not None
            assert result["metrics"] == ["flow_time"]
        finally:
            app.destroy()


class TestExportOnRealData:
    """Acceptance tests for export_pdf() and export_png() against real computed figures."""

    def test_export_single_pdf(self, tmp_path, art_a_data):
        """A single FlowTime figure can be exported to a non-empty PDF file."""
        metric = FlowTimeMetric()
        result = metric.compute(art_a_data, SAFE)
        figs = metric.render(result, SAFE)
        out = tmp_path / "flow_time.pdf"
        export_pdf([figs[0]], out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_export_png(self, tmp_path, art_a_data):
        """A FlowVelocity figure can be exported to a non-empty PNG file."""
        metric = FlowVelocityMetric()
        result = metric.compute(art_a_data, SAFE)
        figs = metric.render(result, SAFE)
        out = tmp_path / "velocity.png"
        export_png(figs[0], out)
        assert out.exists()
        assert out.stat().st_size > 0
