# =============================================================================
# Autor:          Robert Seebauer
# Repository:     https://github.com/Jaegerfeld/situation-report
# KI-Unterstützung: Erstellt mit Unterstützung von Claude (Anthropic)
# Erstellt:       15.04.2026
# Geändert:       18.04.2026
# Lizenz:         BSD-3-Clause (siehe LICENSE)
#
# Fachliche Funktion:
#   Exportiert eine oder mehrere plotly-Figures als PDF-Datei. Jede Figure
#   wird auf einer eigenen Seite ausgegeben. Der Export nutzt kaleido für
#   die Konvertierung zu PNG-Seiten, die anschließend zu einem mehrseitigen
#   PDF zusammengefügt werden. Alternativ kann jede Figure einzeln als PNG
#   exportiert werden.
# =============================================================================

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any

import plotly.graph_objects as go
import plotly.io as pio

if TYPE_CHECKING:
    from .loader import IssueRecord


def export_pdf(figures: list[Any], output_path: Path, width: int = 1400, height: int = 700) -> None:
    """
    Export a list of plotly Figures to a single multi-page PDF file.

    Each figure is rendered as one page. Uses kaleido for rasterization.
    Pages are written sequentially; the PDF is created even if only one
    figure is provided.

    Args:
        figures:     List of plotly Figure objects to export.
        output_path: Destination path for the PDF file (e.g. 'report.pdf').
        width:       Page width in pixels (default 1400).
        height:      Page height in pixels (default 700).

    Raises:
        ValueError:  If figures list is empty.
        ImportError: If kaleido is not installed.
    """
    if not figures:
        raise ValueError("No figures provided for PDF export.")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # kaleido supports direct PDF export for single figures;
    # for multiple pages we write individual PDFs then merge via pypdf if available,
    # otherwise fall back to one PDF per figure.
    if len(figures) == 1:
        pio.write_image(figures[0], str(output_path), format="pdf",
                        width=width, height=height)
        return

    # Try to merge multiple figures into one PDF using pypdf
    try:
        from pypdf import PdfWriter  # optional dependency
        _export_merged_pdf(figures, output_path, width, height)
    except ImportError:
        # Fallback: export one PDF per figure with numeric suffix
        stem = output_path.stem
        for i, fig in enumerate(figures, start=1):
            page_path = output_path.with_name(f"{stem}_page{i}.pdf")
            pio.write_image(fig, str(page_path), format="pdf",
                            width=width, height=height)


def _export_merged_pdf(
    figures: list[Any], output_path: Path, width: int, height: int
) -> None:
    """
    Merge multiple figures into one PDF using pypdf.

    Each figure is written to a temporary single-page PDF, then all pages
    are combined into the final output file.

    Args:
        figures:     List of plotly Figure objects.
        output_path: Final merged PDF path.
        width:       Page width in pixels.
        height:      Page height in pixels.
    """
    import tempfile
    from pypdf import PdfWriter

    writer = PdfWriter()
    tmp_paths: list[Path] = []

    with tempfile.TemporaryDirectory() as tmpdir:
        for i, fig in enumerate(figures):
            tmp_path = Path(tmpdir) / f"page_{i}.pdf"
            pio.write_image(fig, str(tmp_path), format="pdf",
                            width=width, height=height)
            tmp_paths.append(tmp_path)
            writer.append(str(tmp_path))

        with open(output_path, "wb") as f:
            writer.write(f)


def write_zero_day_excel(records: "list[IssueRecord]", path: Path) -> None:
    """
    Write a list of zero-day IssueRecords to an Excel file.

    The output format mirrors IssueTimes.xlsx: one row per issue with the
    same fixed columns. Stage-minute columns are omitted (all were zero or
    irrelevant for these issues). Records are sorted by Project then Key.

    Args:
        records: List of IssueRecord objects excluded as zero-day issues.
        path:    Destination .xlsx path; parent directory is created if needed.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    headers = [
        "Project", "Key", "Issuetype", "Status",
        "Created Date", "Component",
        "First Date", "Implementation Date", "Closed Date",
        "Resolution",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Zero-Day Issues"

    # Header row
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="FFC7CE")  # light red
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    def _fmt(dt: object) -> object:
        """Return datetime as-is (openpyxl formats it); None → empty string."""
        return dt if dt is not None else ""

    for rec in sorted(records, key=lambda r: (r.project, r.key)):
        ws.append([
            rec.project,
            rec.key,
            rec.issuetype,
            rec.status,
            _fmt(rec.created),
            rec.component,
            _fmt(rec.first_date),
            _fmt(rec.implementation_date),
            _fmt(rec.closed_date),
            rec.resolution,
        ])

    # Auto-width for readability
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(path)


def write_report_excel(
    records: "list[IssueRecord]",
    stages: list[str],
    path: Path,
) -> None:
    """
    Write filtered issue records to XLSX in IssueTimes format, extended with
    status group and cycle time columns (Method A and B).

    The output contains all original IssueTimes columns (Project, Key, Issuetype,
    Status, Created Date, Component, First Date, Implementation Date, Closed Date,
    one column per workflow stage, Resolution) followed by three additional columns:
    Status Group (To Do / In Progress / Done), Cycle Time A in calendar days
    (First Date to Closed Date), and Cycle Time B in days (sum of stage minutes
    for all stages except the last, divided by 1440). Cycle time columns are left
    empty for issues that lack First Date or Closed Date.

    Args:
        records: List of IssueRecord objects to export (typically the filtered set).
        stages:  Ordered workflow stage names matching the stage_minutes keys.
        path:    Destination .xlsx path; parent directory is created if needed.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    from .stage_groups import issue_stage_group

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    fixed_headers = [
        "Project", "Key", "Issuetype", "Status",
        "Created Date", "Component",
        "First Date", "Implementation Date", "Closed Date",
    ]
    extra_headers = ["Status Group", "Cycle Time (First->Closed)", "Cycle Time B (days in Status)"]
    headers = fixed_headers + stages + ["Resolution"] + extra_headers

    wb = Workbook()
    ws = wb.active
    ws.title = "Report Data"

    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="BDD7EE")  # light blue
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    def _fmt(dt: object) -> object:
        """Return datetime as-is (openpyxl formats it); None → empty string."""
        return dt if dt is not None else ""

    stages_for_ct_b = stages[:-1] if len(stages) > 1 else stages

    for rec in records:
        if rec.first_date is not None and rec.closed_date is not None:
            ct_a: object = round(
                (rec.closed_date - rec.first_date).total_seconds() / 86400, 2
            )
            minutes = sum(rec.stage_minutes.get(s, 0) for s in stages_for_ct_b)
            ct_b: object = round(minutes / 1440, 2)
        else:
            ct_a = ""
            ct_b = ""

        stage_values = [rec.stage_minutes.get(s, 0) for s in stages]

        ws.append(
            [
                rec.project, rec.key, rec.issuetype, rec.status,
                _fmt(rec.created), rec.component,
                _fmt(rec.first_date), _fmt(rec.implementation_date), _fmt(rec.closed_date),
            ]
            + stage_values
            + [rec.resolution, issue_stage_group(rec), ct_a, ct_b]
        )

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(path)


def export_png(figure: Any, output_path: Path, width: int = 1400, height: int = 700) -> None:
    """
    Export a single plotly Figure as a PNG image.

    Args:
        figure:      A plotly Figure object.
        output_path: Destination path for the PNG file.
        width:       Image width in pixels (default 1400).
        height:      Image height in pixels (default 700).
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pio.write_image(figure, str(output_path), format="png", width=width, height=height)
