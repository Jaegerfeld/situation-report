# =============================================================================
# Autor:          Robert Seebauer
# Repository:     https://github.com/Jaegerfeld/situation-report
# KI-Unterstützung: Erstellt mit Unterstützung von Claude (Anthropic)
# Erstellt:       15.04.2026
# Geändert:       25.04.2026
# Lizenz:         BSD-3-Clause (siehe LICENSE)
#
# Fachliche Funktion:
#   Liest die von transform_data erzeugten XLSX-Dateien (IssueTimes, CFD) ein
#   und wandelt sie in typisierte Datenstrukturen um. Die Stage-Spalten werden
#   dynamisch aus der Kopfzeile der IssueTimes-Datei ermittelt. Fehlende oder
#   leere Datumswerte werden als None behandelt.
# =============================================================================

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

_DT_FMT = "%d.%m.%Y %H:%M:%S"
_DATE_FMT = "%d.%m.%Y"

# Fixed column names in IssueTimes.xlsx that appear before and after stage columns
_FIXED_PREFIX = ["Project", "Group", "Key", "Issuetype", "Status", "Created Date",
                 "Component", "Category", "First Date", "Implementation Date", "Closed Date"]
_FIXED_SUFFIX = ["Resolution"]


def _parse_dt(value: object) -> datetime | None:
    """
    Parse a datetime value from an XLSX cell.

    Accepts datetime objects (openpyxl native) or strings in DD.MM.YYYY HH:MM:SS format.

    Args:
        value: Cell value — datetime, string, or None/empty.

    Returns:
        Parsed datetime, or None if empty or unparseable.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), _DT_FMT)
        except ValueError:
            return None
    return None


def _parse_date(value: object) -> date | None:
    """
    Parse a date value from an XLSX cell.

    Accepts date/datetime objects or strings in DD.MM.YYYY format.

    Args:
        value: Cell value — date, datetime, string, or None/empty.

    Returns:
        Parsed date, or None if empty or unparseable.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), _DATE_FMT).date()
        except ValueError:
            return None
    return None


@dataclass
class IssueRecord:
    """One row from IssueTimes.xlsx — per-issue metadata and stage times."""
    project: str
    key: str
    issuetype: str
    status: str
    created: datetime | None
    component: str
    first_date: datetime | None
    implementation_date: datetime | None
    closed_date: datetime | None
    stage_minutes: dict[str, int]   # stage name -> minutes spent
    resolution: str


@dataclass
class CfdRecord:
    """One row from CFD.xlsx — daily stage counts for the cumulative flow diagram."""
    day: date
    stage_counts: dict[str, int]    # stage name -> number of issues in that stage


@dataclass
class ReportData:
    """
    Container for all data loaded from a transform_data output set.

    Holds issue records, CFD records, the ordered list of workflow stages,
    and the source prefix (e.g. 'ART_A') derived from the file names.
    Optional first_stage / closed_stage mark the system entry and exit
    boundaries from the workflow file (<First> and <Closed> markers).
    """
    issues: list[IssueRecord] = field(default_factory=list)
    cfd: list[CfdRecord] = field(default_factory=list)
    stages: list[str] = field(default_factory=list)
    source_prefix: str = ""
    first_stage: str | None = None   # <First> marker from workflow file
    closed_stage: str | None = None  # <Closed> marker from workflow file


def load_issue_times(path: Path) -> tuple[list[IssueRecord], list[str]]:
    """
    Read an IssueTimes.xlsx file and return issue records plus the ordered stage list.

    Stage columns are inferred dynamically from the header row: all columns
    between 'Closed Date' and 'Resolution' are treated as stage columns.

    Args:
        path: Path to the IssueTimes.xlsx file.

    Returns:
        Tuple of (list of IssueRecord, list of stage names in workflow order).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = iter(ws.iter_rows(values_only=True))
    header = [str(c) if c is not None else "" for c in next(rows)]

    # Determine stage columns dynamically
    try:
        stage_start = header.index("Closed Date") + 1
        stage_end = header.index("Resolution")
    except ValueError as e:
        raise ValueError(f"Unexpected IssueTimes format in {path}: {e}") from e

    stages = header[stage_start:stage_end]

    col = {name: i for i, name in enumerate(header)}

    records: list[IssueRecord] = []
    for row in rows:
        if not any(row):
            continue

        def cell(name: str) -> object:
            return row[col[name]] if name in col else None

        stage_minutes = {
            s: int(row[stage_start + i] or 0)
            for i, s in enumerate(stages)
        }

        records.append(IssueRecord(
            project=str(cell("Project") or ""),
            key=str(cell("Key") or ""),
            issuetype=str(cell("Issuetype") or ""),
            status=str(cell("Status") or ""),
            created=_parse_dt(cell("Created Date")),
            component=str(cell("Component") or ""),
            first_date=_parse_dt(cell("First Date")),
            implementation_date=_parse_dt(cell("Implementation Date")),
            closed_date=_parse_dt(cell("Closed Date")),
            stage_minutes=stage_minutes,
            resolution=str(cell("Resolution") or ""),
        ))

    wb.close()
    return records, stages


def load_cfd(path: Path) -> tuple[list[CfdRecord], list[str]]:
    """
    Read a CFD.xlsx file and return daily stage count records plus the stage list.

    Args:
        path: Path to the CFD.xlsx file.

    Returns:
        Tuple of (list of CfdRecord, list of stage names in workflow order).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = iter(ws.iter_rows(values_only=True))
    header = [str(c) if c is not None else "" for c in next(rows)]
    stages = header[1:]  # first column is "Day"

    records: list[CfdRecord] = []
    for row in rows:
        if not any(row):
            continue
        day = _parse_date(row[0])
        if day is None:
            continue
        stage_counts = {s: int(row[i + 1] or 0) for i, s in enumerate(stages)}
        records.append(CfdRecord(day=day, stage_counts=stage_counts))

    wb.close()
    return records, stages


def _parse_workflow_markers(path: Path) -> tuple[str | None, str | None]:
    """
    Extract <First> and <Closed> stage names from a workflow definition file.

    Only reads lines starting with <First> or <Closed>; all other content is ignored.

    Args:
        path: Path to the workflow .txt file.

    Returns:
        Tuple of (first_stage, closed_stage), each None if the marker is absent.
    """
    first_stage: str | None = None
    closed_stage: str | None = None
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line.startswith("<First>"):
            first_stage = line[7:]
        elif line.startswith("<Closed>"):
            closed_stage = line[8:]
    return first_stage, closed_stage


def load_report_data(
    issue_times_path: Path,
    cfd_path: Path | None = None,
    workflow_path: Path | None = None,
) -> ReportData:
    """
    Load a complete ReportData set from IssueTimes and optionally CFD XLSX files.

    The source_prefix is derived from the IssueTimes filename by stripping
    the '_IssueTimes' suffix (e.g. 'ART_A_IssueTimes.xlsx' → 'ART_A').
    If a workflow file is provided, <First> and <Closed> markers are read
    and stored in first_stage / closed_stage for use by the CFD metric.

    Args:
        issue_times_path: Path to the IssueTimes.xlsx file (required).
        cfd_path:         Path to the CFD.xlsx file (optional).
        workflow_path:    Path to the workflow .txt file (optional).

    Returns:
        Populated ReportData instance.
    """
    issues, stages = load_issue_times(issue_times_path)

    cfd: list[CfdRecord] = []
    if cfd_path is not None:
        cfd, _ = load_cfd(cfd_path)

    first_stage: str | None = None
    closed_stage: str | None = None
    if workflow_path is not None:
        first_stage, closed_stage = _parse_workflow_markers(workflow_path)

    stem = issue_times_path.stem
    prefix = stem.removesuffix("_IssueTimes")

    return ReportData(
        issues=issues,
        cfd=cfd,
        stages=stages,
        source_prefix=prefix,
        first_stage=first_stage,
        closed_stage=closed_stage,
    )
