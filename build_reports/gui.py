# =============================================================================
# Autor:          Robert Seebauer
# Repository:     https://github.com/Jaegerfeld/situation-report
# KI-Unterstützung: Erstellt mit Unterstützung von Claude (Anthropic)
# Erstellt:       16.04.2026
# Geändert:       16.04.2026
# Lizenz:         BSD-3-Clause (siehe LICENSE)
#
# Fachliche Funktion:
#   Grafische Benutzeroberfläche (tkinter) für build_reports. Ermöglicht die
#   Auswahl von IssueTimes- und CFD-XLSX, optionale Filter (Datum, Projekte,
#   Issuetypen), Metrik-Auswahl (Checkboxen) sowie CT-Methode (A/B). Sprache
#   (DE/EN) und Terminologie (SAFe/Global) werden über ein Optionsmenü gewählt.
#   Templates ermöglichen das Speichern und Laden der gesamten Konfiguration
#   als JSON-Datei. Die Berechnung läuft in einem separaten Thread, sodass
#   die Oberfläche reaktionsfähig bleibt.
# =============================================================================

from __future__ import annotations

import json
import tempfile
import threading
import tkinter as tk
import webbrowser
from datetime import date, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk
from typing import Any

from openpyxl import load_workbook
from tkcalendar import Calendar

from .cli import run_reports
from .export import write_zero_day_excel
from .metrics import all_metrics
from .metrics.flow_time import CT_METHOD_A, CT_METHOD_B, FlowTimeMetric
from .terminology import GLOBAL, SAFE, term

# ---------------------------------------------------------------------------
# Language constants
# ---------------------------------------------------------------------------

LANG_DE = "de"
LANG_EN = "en"

# ---------------------------------------------------------------------------
# Translations
# ---------------------------------------------------------------------------

_T: dict[str, dict[str, str]] = {
    LANG_DE: {
        "window_title":      "build_reports",
        "menu_options":      "Optionen",
        "menu_language":     "Sprache",
        "menu_lang_de":      "Deutsch",
        "menu_lang_en":      "English",
        "menu_terminology":  "Terminologie",
        "sec_files":         "Dateien",
        "sec_filter":        "Filter",
        "sec_metrics":       "Metriken",
        "sec_ct":            "CT-Methode (Flow Time)",
        "lbl_issue_times":   "IssueTimes",
        "lbl_cfd":           "CFD (optional)",
        "lbl_from":          "Von (YYYY-MM-DD)",
        "lbl_to":            "Bis (YYYY-MM-DD)",
        "lbl_projects":      "Projekte",
        "lbl_issuetypes":    "Issuetypen",
        "btn_browse":        "Durchsuchen\u2026",
        "btn_pick":          "\u25be",
        "btn_all":           "Alle",
        "btn_none":          "Keine",
        "btn_show":          "Anzeigen im Browser",
        "btn_pdf":           "Als PDF speichern",
        "hint_csv":          "kommagetrennt",
        "lbl_log":           "Log",
        "ct_a":              "A \u2013 Kalendertage (Closed Date \u2212 First Date)",
        "ct_b":              "B \u2013 Stage-Minuten (First bis Closed, exkl.)",
        "dlg_issue_times":   "IssueTimes-Datei w\u00e4hlen",
        "dlg_cfd":           "CFD-Datei w\u00e4hlen",
        "dlg_pdf":           "PDF speichern unter",
        "dlg_pick_date":     "Datum w\u00e4hlen",
        "dlg_projects":      "Projekte w\u00e4hlen",
        "dlg_issuetypes":    "Issuetypen w\u00e4hlen",
        "btn_cal":           "\U0001f4c5",
        "btn_ok":            "OK",
        "btn_last_365":      "Letzte 365 Tage",
        "err_no_file":       "FEHLER: Keine IssueTimes-Datei ausgew\u00e4hlt.",
        "err_not_found":     "FEHLER: Datei nicht gefunden: {}",
        "err_from_date":     "FEHLER: Ung\u00fcltiges Von-Datum '{}' (erwartet YYYY-MM-DD).",
        "err_to_date":       "FEHLER: Ung\u00fcltiges Bis-Datum '{}' (erwartet YYYY-MM-DD).",
        "log_check_ok":      "Stages: IssueTimes und CFD stimmen \u00fcberein.",
        "log_check_miss_cfd":"  Stage nur in IssueTimes: {}",
        "log_check_miss_it": "  Stage nur in CFD: {}",
        "log_started":       "--- Berechnung gestartet ---",
        "log_done":          "--- Fertig ---",
        "log_pdf_started":   "--- PDF-Export gestartet ---",
        "log_no_figs":       "Keine Diagramme erzeugt.",
        "log_figs_opened":   "  \u2192 {} Diagramm(e) im Browser ge\u00f6ffnet.",
        "log_unk_metric":    "  WARNING: Unbekannte Metrik '{}' \u2014 \u00fcbersprungen.",
        "log_error":         "FEHLER: {}",
        "menu_template":     "Templates",
        "menu_tpl_save":     "Speichern\u2026",
        "menu_tpl_load":     "Laden\u2026",
        "dlg_tpl_save":      "Template speichern",
        "dlg_tpl_load":      "Template laden",
        "log_tpl_saved":     "Template gespeichert: {}",
        "log_tpl_loaded":    "Template geladen: {}",
        "log_tpl_error":     "Fehler beim Template: {}",
        "log_zero_day_xlsx": "  {} Zero-Day Issue(s) exportiert: {}",
        "log_zero_day_log":  "  Zero-Day Issues ({}): {}",
        # Tooltips
        "tip_issue_times":   "IssueTimes.xlsx aus transform_data \u2014 enth\u00e4lt alle Issues mit Datums- und Stufenangaben.",
        "tip_cfd":           "CFD.xlsx aus transform_data \u2014 optional, wird nur f\u00fcr die CFD-Metrik ben\u00f6tigt.",
        "tip_browse":        "Datei ausw\u00e4hlen \u2026",
        "tip_from":          "Nur Issues einbeziehen, die ab diesem Datum abgeschlossen wurden (inkl.).",
        "tip_to":            "Nur Issues einbeziehen, die bis zu diesem Datum abgeschlossen wurden (inkl.).",
        "tip_cal":           "Kalender \u00f6ffnen",
        "tip_last_365":      "Setzt Von und Bis auf die letzten 365 Tage bis heute.",
        "tip_projects":      "Kommagetrennte Projektschl\u00fcssel, z.\u202fB. \u201eARTA, ARTB\u201c. Leer = alle Projekte.",
        "tip_issuetypes":    "Kommagetrennte Issuetypen, z.\u202fB. \u201eFeature, Bug\u201c. Leer = alle Typen.",
        "tip_pick":          "Aus der geladenen IssueTimes-Datei ausw\u00e4hlen.",
        "tip_ct_a":          "Berechnet CT als Differenz der Kalendertage zwischen First Date und Closed Date.",
        "tip_ct_b":          "Berechnet CT als Summe der Stage-Minuten von First Date bis Closed Date (letzte Stage ausgeschlossen).",
        "tip_show":          "Metriken berechnen und Ergebnisse im Standard-Browser anzeigen.",
        "tip_pdf":           "Metriken berechnen und alle Diagramme als PDF-Datei exportieren.",
        "tip_metric_flow_time":         "Wie lange braucht ein Issue von Start bis Abschluss?",
        "tip_metric_flow_velocity":     "Wie viele Issues werden pro Zeiteinheit abgeschlossen?",
        "tip_metric_flow_load":         "Wie viele Issues befinden sich gleichzeitig in Bearbeitung?",
        "tip_metric_cfd":               "Kumulative Anzahl von Issues pro Stage \u00fcber die Zeit.",
        "tip_metric_flow_distribution": "Verteilung der Issues nach Typ oder Kategorie.",
    },
    LANG_EN: {
        "window_title":      "build_reports",
        "menu_options":      "Options",
        "menu_language":     "Language",
        "menu_lang_de":      "Deutsch",
        "menu_lang_en":      "English",
        "menu_terminology":  "Terminology",
        "sec_files":         "Files",
        "sec_filter":        "Filter",
        "sec_metrics":       "Metrics",
        "sec_ct":            "CT Method (Flow Time)",
        "lbl_issue_times":   "IssueTimes",
        "lbl_cfd":           "CFD (optional)",
        "lbl_from":          "From (YYYY-MM-DD)",
        "lbl_to":            "To (YYYY-MM-DD)",
        "lbl_projects":      "Projects",
        "lbl_issuetypes":    "Issue Types",
        "btn_browse":        "Browse\u2026",
        "btn_pick":          "\u25be",
        "btn_all":           "All",
        "btn_none":          "None",
        "btn_show":          "Show in Browser",
        "btn_pdf":           "Save as PDF",
        "hint_csv":          "comma-separated",
        "lbl_log":           "Log",
        "ct_a":              "A \u2013 Calendar days (Closed Date \u2212 First Date)",
        "ct_b":              "B \u2013 Stage minutes (First to Closed, excl.)",
        "dlg_issue_times":   "Select IssueTimes file",
        "dlg_cfd":           "Select CFD file",
        "dlg_pdf":           "Save PDF as",
        "dlg_pick_date":     "Pick Date",
        "dlg_projects":      "Select Projects",
        "dlg_issuetypes":    "Select Issue Types",
        "btn_cal":           "\U0001f4c5",
        "btn_ok":            "OK",
        "btn_last_365":      "Last 365 Days",
        "err_no_file":       "ERROR: No IssueTimes file selected.",
        "err_not_found":     "ERROR: File not found: {}",
        "err_from_date":     "ERROR: Invalid From date '{}' (expected YYYY-MM-DD).",
        "err_to_date":       "ERROR: Invalid To date '{}' (expected YYYY-MM-DD).",
        "log_check_ok":      "Stages: IssueTimes and CFD are consistent.",
        "log_check_miss_cfd":"  Stage only in IssueTimes: {}",
        "log_check_miss_it": "  Stage only in CFD: {}",
        "log_started":       "--- Computation started ---",
        "log_done":          "--- Done ---",
        "log_pdf_started":   "--- PDF export started ---",
        "log_no_figs":       "No figures produced.",
        "log_figs_opened":   "  \u2192 {} figure(s) opened in browser.",
        "log_unk_metric":    "  WARNING: Unknown metric '{}' \u2014 skipped.",
        "log_error":         "ERROR: {}",
        "menu_template":     "Templates",
        "menu_tpl_save":     "Save\u2026",
        "menu_tpl_load":     "Load\u2026",
        "dlg_tpl_save":      "Save Template",
        "dlg_tpl_load":      "Load Template",
        "log_tpl_saved":     "Template saved: {}",
        "log_tpl_loaded":    "Template loaded: {}",
        "log_tpl_error":     "Template error: {}",
        "log_zero_day_xlsx": "  {} zero-day issue(s) exported: {}",
        "log_zero_day_log":  "  Zero-Day Issues ({}): {}",
        # Tooltips
        "tip_issue_times":   "IssueTimes.xlsx from transform_data \u2014 contains all issues with dates and stage data.",
        "tip_cfd":           "CFD.xlsx from transform_data \u2014 optional, only required for the CFD metric.",
        "tip_browse":        "Select file \u2026",
        "tip_from":          "Include only issues closed on or after this date (inclusive).",
        "tip_to":            "Include only issues closed on or before this date (inclusive).",
        "tip_cal":           "Open calendar picker",
        "tip_last_365":      "Sets From and To to the last 365 days ending today.",
        "tip_projects":      "Comma-separated project keys, e.g. \u201cARTA, ARTB\u201d. Empty = all projects.",
        "tip_issuetypes":    "Comma-separated issue types, e.g. \u201cFeature, Bug\u201d. Empty = all types.",
        "tip_pick":          "Select from the loaded IssueTimes file.",
        "tip_ct_a":          "Computes CT as the calendar-day difference between First Date and Closed Date.",
        "tip_ct_b":          "Computes CT as the sum of stage minutes from First Date to Closed Date (last stage excluded).",
        "tip_show":          "Compute metrics and display results in the default browser.",
        "tip_pdf":           "Compute metrics and export all charts as a PDF file.",
        "tip_metric_flow_time":         "How long does an issue take from start to close?",
        "tip_metric_flow_velocity":     "How many issues are completed per time unit?",
        "tip_metric_flow_load":         "How many issues are simultaneously in progress?",
        "tip_metric_cfd":               "Cumulative count of issues per stage over time.",
        "tip_metric_flow_distribution": "Distribution of issues by type or category.",
    },
}


# ---------------------------------------------------------------------------
# Module-level helpers (testable without a running display)
# ---------------------------------------------------------------------------

def _read_available_filters(path: Path) -> tuple[list[str], list[str]]:
    """
    Read unique project keys and issue types from an IssueTimes.xlsx file.

    Only the Project and Issuetype columns are read; all other columns are
    skipped for efficiency.

    Args:
        path: Path to the IssueTimes.xlsx file.

    Returns:
        Tuple of (sorted project list, sorted issuetype list).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = iter(ws.iter_rows(values_only=True))
    header = [str(c) if c is not None else "" for c in next(rows)]
    col = {name: i for i, name in enumerate(header)}

    projects: set[str] = set()
    issuetypes: set[str] = set()
    for row in rows:
        if not any(row):
            continue
        if "Project" in col and row[col["Project"]]:
            projects.add(str(row[col["Project"]]))
        if "Issuetype" in col and row[col["Issuetype"]]:
            issuetypes.add(str(row[col["Issuetype"]]))

    wb.close()
    return sorted(projects), sorted(issuetypes)


def _check_stage_consistency(
    issue_times_path: Path, cfd_path: Path
) -> tuple[list[str], list[str]]:
    """
    Compare stage columns between an IssueTimes and a CFD XLSX file.

    Reads only the header rows for efficiency.

    Args:
        issue_times_path: Path to IssueTimes.xlsx.
        cfd_path:         Path to CFD.xlsx.

    Returns:
        Tuple of (only_in_issue_times, only_in_cfd): sorted lists of stage names
        that appear in one file but not the other. Both empty means consistent.

    Raises:
        ValueError: If IssueTimes is missing the expected Closed Date / Resolution columns.
    """
    # Read IssueTimes header
    wb = load_workbook(issue_times_path, read_only=True, data_only=True)
    header_it = [str(c) if c is not None else "" for c in
                 next(wb.active.iter_rows(values_only=True))]
    wb.close()

    try:
        stage_start = header_it.index("Closed Date") + 1
        stage_end = header_it.index("Resolution")
    except ValueError as exc:
        raise ValueError(
            f"Unexpected IssueTimes format in {issue_times_path}: {exc}"
        ) from exc
    it_stages = set(header_it[stage_start:stage_end])

    # Read CFD header
    wb = load_workbook(cfd_path, read_only=True, data_only=True)
    header_cfd = [str(c) if c is not None else "" for c in
                  next(wb.active.iter_rows(values_only=True))]
    wb.close()
    cfd_stages = set(header_cfd[1:])  # skip "Day" column

    only_in_it = sorted(it_stages - cfd_stages)
    only_in_cfd = sorted(cfd_stages - it_stages)
    return only_in_it, only_in_cfd


def _default_date_range() -> tuple[date, date]:
    """
    Return the default filter date range: today minus 365 days through today.

    Returns:
        Tuple of (from_date, to_date) covering the last 365 days inclusive.
    """
    today = date.today()
    return today - timedelta(days=365), today


def _parse_date_safe(value: str) -> date | None:
    """
    Parse a date string in YYYY-MM-DD format, returning None on empty input.

    Args:
        value: Date string; empty string returns None.

    Returns:
        Parsed date or None.

    Raises:
        ValueError: If the string is non-empty but not a valid ISO date.
    """
    stripped = value.strip()
    if not stripped:
        return None
    return date.fromisoformat(stripped)


def _split_csv(value: str) -> list[str]:
    """
    Split a comma-separated string into a list of non-empty, stripped items.

    Args:
        value: Comma-separated string from a GUI entry widget.

    Returns:
        List of non-empty strings.
    """
    return [item.strip() for item in value.split(",") if item.strip()]


def _build_combined_html(
    figures: list,
    section_breaks: "dict[int, str] | None" = None,
) -> str:
    """
    Combine multiple plotly Figure objects into a single self-contained HTML string.

    The first figure includes the full Plotly.js CDN bundle; subsequent figures
    reuse the already-loaded library to keep file size reasonable.

    An optional ``section_breaks`` dict maps figure index → heading text. When
    a heading is present for index ``i``, an ``<h2>`` element is inserted before
    that figure's div — useful for labelling metric groups in the browser view.

    Args:
        figures:        List of plotly Figure objects to embed.
        section_breaks: Optional dict mapping figure index to a heading string.

    Returns:
        Complete HTML document as a string.
    """
    breaks = section_breaks or {}
    parts = []
    for i, fig in enumerate(figures):
        if i in breaks:
            parts.append(
                f'<h2 class="metric-heading">{breaks[i]}</h2>'
            )
        parts.append(
            fig.to_html(
                include_plotlyjs="cdn" if i == 0 else False,
                full_html=False,
                config={"responsive": True},
            )
        )
    body = "\n".join(parts)
    return (
        "<!DOCTYPE html><html><head>"
        "<meta charset='utf-8'>"
        "<style>"
        "body{margin:16px;font-family:sans-serif;background:#fff;}"
        "div.plotly-graph-div{margin-bottom:32px;}"
        "h2.metric-heading{"
        "font-size:1.5rem;font-weight:700;margin:32px 0 4px 0;"
        "padding-bottom:4px;border-bottom:2px solid #d0d0d0;}"
        "</style>"
        "</head>"
        f"<body>{body}</body></html>"
    )


# ---------------------------------------------------------------------------
# Tooltip helper
# ---------------------------------------------------------------------------

class _ToolTip:
    """
    Lightweight hover tooltip for any tkinter widget.

    Binds <Enter> and <Leave> events on creation. The displayed text can be
    updated at any time via update_text() to support runtime language switches.

    Args:
        widget: The widget to attach the tooltip to.
        text:   Initial tooltip text.
    """

    def __init__(self, widget: "tk.Widget", text: str) -> None:
        self._widget = widget
        self._text = text
        self._tip_window: "tk.Toplevel | None" = None
        widget.bind("<Enter>", self._show, add="+")
        widget.bind("<Leave>", self._hide, add="+")

    def update_text(self, text: str) -> None:
        """Replace the tooltip text (takes effect on the next hover)."""
        self._text = text

    def _show(self, _event: object = None) -> None:
        """Display the tooltip window near the widget."""
        if self._tip_window or not self._text:
            return
        x = self._widget.winfo_rootx() + 20
        y = self._widget.winfo_rooty() + self._widget.winfo_height() + 4
        self._tip_window = tw = tk.Toplevel(self._widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tk.Label(
            tw,
            text=self._text,
            background="#ffffe0",
            relief="solid",
            borderwidth=1,
            font=("", 9),
            wraplength=300,
            justify="left",
            padx=5,
            pady=3,
        ).pack()

    def _hide(self, _event: object = None) -> None:
        """Destroy the tooltip window."""
        if self._tip_window:
            self._tip_window.destroy()
            self._tip_window = None


# Template version bump when the schema changes in a backward-incompatible way.
_TEMPLATE_VERSION = 1


def _build_template_dict(
    issue_times: str,
    cfd: str,
    from_date: str,
    to_date: str,
    projects: str,
    issuetypes: str,
    terminology: str,
    ct_method: str,
    metrics: dict[str, bool],
    language: str,
) -> dict:
    """
    Assemble the template dictionary that is written to JSON.

    All string values are stored as-is; file paths are kept as strings so the
    template remains portable (the caller decides whether to resolve them).

    Args:
        issue_times:  Absolute path string for IssueTimes.xlsx (may be empty).
        cfd:          Absolute path string for CFD.xlsx (may be empty).
        from_date:    ISO date string (YYYY-MM-DD) or empty.
        to_date:      ISO date string (YYYY-MM-DD) or empty.
        projects:     Comma-separated project keys (may be empty).
        issuetypes:   Comma-separated issue types (may be empty).
        terminology:  Terminology mode constant (SAFE or GLOBAL).
        ct_method:    CT calculation method constant (CT_METHOD_A or CT_METHOD_B).
        metrics:      Dict mapping metric_id → bool (True = selected).
        language:     Language code (LANG_DE or LANG_EN).

    Returns:
        JSON-serialisable dict with a ``version`` key.
    """
    return {
        "version": _TEMPLATE_VERSION,
        "issue_times": issue_times,
        "cfd": cfd,
        "from_date": from_date,
        "to_date": to_date,
        "projects": projects,
        "issuetypes": issuetypes,
        "terminology": terminology,
        "ct_method": ct_method,
        "metrics": metrics,
        "language": language,
    }


def _parse_template_dict(data: dict) -> dict:
    """
    Validate and normalise a template dict loaded from JSON.

    Unknown keys are ignored; missing keys fall back to empty/default values so
    that templates created by older versions remain loadable.

    Args:
        data: Raw dict parsed from a JSON template file.

    Returns:
        Normalised dict with guaranteed keys matching _build_template_dict output.

    Raises:
        ValueError: If ``data`` is not a dict or has an incompatible version.
    """
    if not isinstance(data, dict):
        raise ValueError("Template file does not contain a JSON object.")
    version = data.get("version", 1)
    if not isinstance(version, int) or version > _TEMPLATE_VERSION:
        raise ValueError(
            f"Unsupported template version {version} "
            f"(this build supports up to {_TEMPLATE_VERSION})."
        )
    return {
        "version": version,
        "issue_times": str(data.get("issue_times", "")),
        "cfd": str(data.get("cfd", "")),
        "from_date": str(data.get("from_date", "")),
        "to_date": str(data.get("to_date", "")),
        "projects": str(data.get("projects", "")),
        "issuetypes": str(data.get("issuetypes", "")),
        "terminology": str(data.get("terminology", SAFE)),
        "ct_method": str(data.get("ct_method", CT_METHOD_A)),
        "metrics": dict(data.get("metrics", {})),
        "language": str(data.get("language", LANG_DE)),
    }


# ---------------------------------------------------------------------------
# Main application
# ---------------------------------------------------------------------------

class BuildReportsApp(tk.Tk):
    """
    Main tkinter application window for build_reports.

    Provides file selection, filter controls, metric checkboxes, CT method
    toggle, a browser preview button, and PDF export. Language (DE/EN) and
    terminology (SAFe/Global) are controlled via the Options menu. Computation
    runs in a background thread to keep the UI responsive.
    """

    def __init__(self) -> None:
        super().__init__()
        self.resizable(True, True)
        self.minsize(520, 600)

        self._plugins = all_metrics()

        # --- State variables ---
        self._lang_var = tk.StringVar(value=LANG_DE)
        self._issue_times_var = tk.StringVar()
        self._cfd_var = tk.StringVar()
        _from_default, _to_default = _default_date_range()
        self._from_date_var = tk.StringVar(value=str(_from_default))
        self._to_date_var = tk.StringVar(value=str(_to_default))
        self._projects_var = tk.StringVar()
        self._issuetypes_var = tk.StringVar()
        self._terminology_var = tk.StringVar(value=SAFE)
        self._ct_method_var = tk.StringVar(value=CT_METHOD_A)
        self._available_projects: list[str] = []
        self._available_issuetypes: list[str] = []
        self._metric_vars: dict[str, tk.BooleanVar] = {
            p.metric_id: tk.BooleanVar(value=True) for p in self._plugins
        }

        # Translatable widget references: list of (widget, tr_key)
        self._i18n: list[tuple[Any, str]] = []
        # Metric checkbuttons for terminology-driven label updates
        self._metric_checkbuttons: list[tuple[tk.Checkbutton, str]] = []
        # Tooltips: list of (_ToolTip, tr_key)
        self._tips: list[tuple[_ToolTip, str]] = []

        # React to language / terminology changes
        self._lang_var.trace_add("write", lambda *_: self._apply_language())
        self._terminology_var.trace_add("write", lambda *_: self._apply_terminology())

        self._build_menubar()
        self._build_ui()
        self._apply_language()  # set initial titles / labels

    # -------------------------------------------------------------------------
    # Translation helper
    # -------------------------------------------------------------------------

    def _tr(self, key: str) -> str:
        """
        Return the translated string for the current language.

        Args:
            key: Translation key defined in _T.

        Returns:
            Translated string, or key itself as fallback.
        """
        return _T.get(self._lang_var.get(), _T[LANG_DE]).get(key, key)

    # -------------------------------------------------------------------------
    # Menu bar
    # -------------------------------------------------------------------------

    def _build_menubar(self) -> None:
        """Build (or rebuild) the top menu bar with Options → Language + Terminology."""
        menubar = tk.Menu(self)

        options_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=self._tr("menu_options"), menu=options_menu)

        # Language submenu
        lang_menu = tk.Menu(options_menu, tearoff=0)
        options_menu.add_cascade(label=self._tr("menu_language"), menu=lang_menu)
        lang_menu.add_radiobutton(
            label=self._tr("menu_lang_de"),
            variable=self._lang_var,
            value=LANG_DE,
        )
        lang_menu.add_radiobutton(
            label=self._tr("menu_lang_en"),
            variable=self._lang_var,
            value=LANG_EN,
        )

        # Terminology submenu
        options_menu.add_separator()
        options_menu.add_cascade(
            label=self._tr("menu_terminology"),
            menu=self._build_terminology_menu(options_menu),
        )

        # Templates menu
        tpl_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=self._tr("menu_template"), menu=tpl_menu)
        tpl_menu.add_command(label=self._tr("menu_tpl_save"), command=self._save_template)
        tpl_menu.add_command(label=self._tr("menu_tpl_load"), command=self._load_template)

        self.config(menu=menubar)

    def _build_terminology_menu(self, parent: tk.Menu) -> tk.Menu:
        """Create and return the Terminology sub-menu."""
        term_menu = tk.Menu(parent, tearoff=0)
        term_menu.add_radiobutton(label="SAFe",   variable=self._terminology_var, value=SAFE)
        term_menu.add_radiobutton(label="Global", variable=self._terminology_var, value=GLOBAL)
        return term_menu

    # -------------------------------------------------------------------------
    # UI construction
    # -------------------------------------------------------------------------

    def _build_ui(self) -> None:
        """Build and arrange all main window widgets."""
        pad = {"padx": 8, "pady": 3}
        self.columnconfigure(1, weight=1)
        row = 0

        # --- Files ---
        row = self._section_header("sec_files", row)

        lbl = tk.Label(self, text=self._tr("lbl_issue_times"), anchor="w")
        lbl.grid(row=row, column=0, sticky="w", **pad)
        self._i18n.append((lbl, "lbl_issue_times"))

        it_entry = tk.Entry(self, textvariable=self._issue_times_var, state="readonly", width=50)
        it_entry.grid(row=row, column=1, sticky="ew", **pad)
        self._tips.append((_ToolTip(it_entry, self._tr("tip_issue_times")), "tip_issue_times"))

        btn = ttk.Button(self, text=self._tr("btn_browse"), command=self._pick_issue_times)
        btn.grid(row=row, column=2, **pad)
        self._i18n.append((btn, "btn_browse"))
        self._tips.append((_ToolTip(btn, self._tr("tip_browse")), "tip_browse"))
        row += 1

        lbl = tk.Label(self, text=self._tr("lbl_cfd"), anchor="w")
        lbl.grid(row=row, column=0, sticky="w", **pad)
        self._i18n.append((lbl, "lbl_cfd"))

        cfd_entry = tk.Entry(self, textvariable=self._cfd_var, state="readonly", width=50)
        cfd_entry.grid(row=row, column=1, sticky="ew", **pad)
        self._tips.append((_ToolTip(cfd_entry, self._tr("tip_cfd")), "tip_cfd"))

        btn = ttk.Button(self, text=self._tr("btn_browse"), command=self._pick_cfd)
        btn.grid(row=row, column=2, **pad)
        self._i18n.append((btn, "btn_browse"))
        self._tips.append((_ToolTip(btn, self._tr("tip_browse")), "tip_browse"))
        row += 1

        # --- Filter ---
        row = self._section_header("sec_filter", row)

        lbl = tk.Label(self, text=self._tr("lbl_from"), anchor="w")
        lbl.grid(row=row, column=0, sticky="w", **pad)
        self._i18n.append((lbl, "lbl_from"))
        from_entry = tk.Entry(self, textvariable=self._from_date_var, width=20)
        from_entry.grid(row=row, column=1, sticky="w", **pad)
        self._tips.append((_ToolTip(from_entry, self._tr("tip_from")), "tip_from"))
        btn = ttk.Button(self, text=self._tr("btn_cal"), width=3,
                         command=lambda: self._pick_date(self._from_date_var))
        btn.grid(row=row, column=2, sticky="w", **pad)
        self._i18n.append((btn, "btn_cal"))
        self._tips.append((_ToolTip(btn, self._tr("tip_cal")), "tip_cal"))
        row += 1

        lbl = tk.Label(self, text=self._tr("lbl_to"), anchor="w")
        lbl.grid(row=row, column=0, sticky="w", **pad)
        self._i18n.append((lbl, "lbl_to"))
        to_entry = tk.Entry(self, textvariable=self._to_date_var, width=20)
        to_entry.grid(row=row, column=1, sticky="w", **pad)
        self._tips.append((_ToolTip(to_entry, self._tr("tip_to")), "tip_to"))
        btn = ttk.Button(self, text=self._tr("btn_cal"), width=3,
                         command=lambda: self._pick_date(self._to_date_var))
        btn.grid(row=row, column=2, sticky="w", **pad)
        self._i18n.append((btn, "btn_cal"))
        self._tips.append((_ToolTip(btn, self._tr("tip_cal")), "tip_cal"))
        row += 1

        btn = ttk.Button(self, text=self._tr("btn_last_365"),
                         command=self._set_last_365_days)
        btn.grid(row=row, column=0, columnspan=2, sticky="w", padx=16, pady=2)
        self._i18n.append((btn, "btn_last_365"))
        self._tips.append((_ToolTip(btn, self._tr("tip_last_365")), "tip_last_365"))
        row += 1

        lbl = tk.Label(self, text=self._tr("lbl_projects"), anchor="w")
        lbl.grid(row=row, column=0, sticky="w", **pad)
        self._i18n.append((lbl, "lbl_projects"))
        proj_entry = tk.Entry(self, textvariable=self._projects_var, width=44)
        proj_entry.grid(row=row, column=1, sticky="ew", **pad)
        self._tips.append((_ToolTip(proj_entry, self._tr("tip_projects")), "tip_projects"))
        btn = ttk.Button(self, text=self._tr("btn_pick"), width=3,
                         command=lambda: self._open_multiselect(
                             self._projects_var, self._available_projects, "dlg_projects"))
        btn.grid(row=row, column=2, sticky="w", **pad)
        self._i18n.append((btn, "btn_pick"))
        self._tips.append((_ToolTip(btn, self._tr("tip_pick")), "tip_pick"))
        row += 1

        lbl = tk.Label(self, text=self._tr("lbl_issuetypes"), anchor="w")
        lbl.grid(row=row, column=0, sticky="w", **pad)
        self._i18n.append((lbl, "lbl_issuetypes"))
        it_type_entry = tk.Entry(self, textvariable=self._issuetypes_var, width=44)
        it_type_entry.grid(row=row, column=1, sticky="ew", **pad)
        self._tips.append((_ToolTip(it_type_entry, self._tr("tip_issuetypes")), "tip_issuetypes"))
        btn = ttk.Button(self, text=self._tr("btn_pick"), width=3,
                         command=lambda: self._open_multiselect(
                             self._issuetypes_var, self._available_issuetypes, "dlg_issuetypes"))
        btn.grid(row=row, column=2, sticky="w", **pad)
        self._i18n.append((btn, "btn_pick"))
        self._tips.append((_ToolTip(btn, self._tr("tip_pick")), "tip_pick"))
        row += 1

        # --- Metrics ---
        row = self._section_header("sec_metrics", row)

        self._metric_checkbuttons.clear()
        for plugin in self._plugins:
            cb = tk.Checkbutton(
                self,
                text=term(plugin.metric_id, self._terminology_var.get()),
                variable=self._metric_vars[plugin.metric_id],
                anchor="w",
            )
            cb.grid(row=row, column=0, columnspan=2, sticky="w", padx=16, pady=1)
            self._metric_checkbuttons.append((cb, plugin.metric_id))
            tip_key = f"tip_metric_{plugin.metric_id}"
            if tip_key in _T[LANG_DE]:
                self._tips.append((_ToolTip(cb, self._tr(tip_key)), tip_key))
            row += 1

        btn_frame = tk.Frame(self)
        btn_frame.grid(row=row, column=0, columnspan=3, sticky="w", padx=16, pady=2)
        btn_all = ttk.Button(btn_frame, text=self._tr("btn_all"), width=6,
                             command=self._select_all_metrics)
        btn_all.pack(side="left", padx=2)
        self._i18n.append((btn_all, "btn_all"))
        btn_none = ttk.Button(btn_frame, text=self._tr("btn_none"), width=6,
                              command=self._deselect_all_metrics)
        btn_none.pack(side="left", padx=2)
        self._i18n.append((btn_none, "btn_none"))
        row += 1

        # --- CT Method ---
        row = self._section_header("sec_ct", row)

        ct_frame = tk.Frame(self)
        ct_frame.grid(row=row, column=0, columnspan=3, sticky="w", padx=16, pady=4)
        for value, key, tip_key in [
            (CT_METHOD_A, "ct_a", "tip_ct_a"),
            (CT_METHOD_B, "ct_b", "tip_ct_b"),
        ]:
            rb = tk.Radiobutton(
                ct_frame,
                text=self._tr(key),
                variable=self._ct_method_var,
                value=value,
            )
            rb.pack(anchor="w", padx=8)
            self._i18n.append((rb, key))
            self._tips.append((_ToolTip(rb, self._tr(tip_key)), tip_key))
        row += 1

        # --- Action buttons ---
        ttk.Separator(self, orient="horizontal").grid(
            row=row, column=0, columnspan=3, sticky="ew", pady=6
        )
        row += 1

        action_frame = tk.Frame(self)
        action_frame.grid(row=row, column=0, columnspan=3, pady=4)
        self._show_btn = ttk.Button(
            action_frame, text=self._tr("btn_show"), command=self._run_show
        )
        self._show_btn.pack(side="left", padx=8)
        self._i18n.append((self._show_btn, "btn_show"))
        self._tips.append((_ToolTip(self._show_btn, self._tr("tip_show")), "tip_show"))
        self._pdf_btn = ttk.Button(
            action_frame, text=self._tr("btn_pdf"), command=self._run_export_pdf
        )
        self._pdf_btn.pack(side="left", padx=8)
        self._i18n.append((self._pdf_btn, "btn_pdf"))
        self._tips.append((_ToolTip(self._pdf_btn, self._tr("tip_pdf")), "tip_pdf"))
        row += 1

        # --- Log area ---
        ttk.Separator(self, orient="horizontal").grid(
            row=row, column=0, columnspan=3, sticky="ew", pady=4
        )
        row += 1

        lbl = tk.Label(self, text=self._tr("lbl_log"), anchor="w")
        lbl.grid(row=row, column=0, sticky="w", **pad)
        self._i18n.append((lbl, "lbl_log"))
        row += 1

        self._log_area = scrolledtext.ScrolledText(
            self, height=10, state="disabled", wrap="word"
        )
        self._log_area.grid(
            row=row, column=0, columnspan=3, sticky="nsew", **pad
        )
        self.rowconfigure(row, weight=1)

    def _section_header(self, tr_key: str, row: int) -> int:
        """Insert a labelled separator, store the Label for i18n, and return next row."""
        ttk.Separator(self, orient="horizontal").grid(
            row=row, column=0, columnspan=3, sticky="ew", pady=(8, 2)
        )
        row += 1
        lbl = tk.Label(self, text=self._tr(tr_key), font=("", 9, "bold"), anchor="w")
        lbl.grid(row=row, column=0, columnspan=3, sticky="w", padx=8)
        self._i18n.append((lbl, tr_key))
        row += 1
        return row

    # -------------------------------------------------------------------------
    # Language / terminology updates
    # -------------------------------------------------------------------------

    def _apply_language(self) -> None:
        """Update all translatable widgets, tooltips, and the window title."""
        self.title(self._tr("window_title"))
        for widget, key in self._i18n:
            widget.config(text=self._tr(key))
        for tip, key in self._tips:
            tip.update_text(self._tr(key))
        self._build_menubar()

    def _apply_terminology(self) -> None:
        """Update metric checkbutton labels to reflect the current terminology."""
        current_term = self._terminology_var.get()
        for cb, metric_id in self._metric_checkbuttons:
            cb.config(text=term(metric_id, current_term))

    # -------------------------------------------------------------------------
    # Template save / load
    # -------------------------------------------------------------------------

    def _save_template(self) -> None:
        """
        Collect the current UI state and write it as a JSON template file.

        Opens a save-file dialog so the user can choose the destination. The
        template includes all file paths, filter values, metric selection,
        CT method, terminology, and language setting.
        """
        path = filedialog.asksaveasfilename(
            title=self._tr("dlg_tpl_save"),
            defaultextension=".json",
            filetypes=[("JSON-Dateien", "*.json"), ("Alle Dateien", "*.*")],
        )
        if not path:
            return
        try:
            tpl = _build_template_dict(
                issue_times=self._issue_times_var.get(),
                cfd=self._cfd_var.get(),
                from_date=self._from_date_var.get(),
                to_date=self._to_date_var.get(),
                projects=self._projects_var.get(),
                issuetypes=self._issuetypes_var.get(),
                terminology=self._terminology_var.get(),
                ct_method=self._ct_method_var.get(),
                metrics={mid: var.get() for mid, var in self._metric_vars.items()},
                language=self._lang_var.get(),
            )
            Path(path).write_text(
                json.dumps(tpl, indent=2, ensure_ascii=False), encoding="utf-8"
            )
            self._log(self._tr("log_tpl_saved").format(Path(path).name))
        except Exception as exc:
            self._log(self._tr("log_tpl_error").format(exc))
            messagebox.showerror(self._tr("menu_template"), str(exc))

    def _load_template(self) -> None:
        """
        Open a JSON template file and apply its settings to the UI.

        Paths stored in the template are applied directly; if a stored file path
        no longer exists a warning is logged but loading continues. The language
        is applied last so all label updates reflect the loaded language.
        """
        path = filedialog.askopenfilename(
            title=self._tr("dlg_tpl_load"),
            filetypes=[("JSON-Dateien", "*.json"), ("Alle Dateien", "*.*")],
        )
        if not path:
            return
        try:
            raw = json.loads(Path(path).read_text(encoding="utf-8"))
            state = _parse_template_dict(raw)
        except Exception as exc:
            self._log(self._tr("log_tpl_error").format(exc))
            messagebox.showerror(self._tr("menu_template"), str(exc))
            return

        self._issue_times_var.set(state["issue_times"])
        self._cfd_var.set(state["cfd"])
        self._from_date_var.set(state["from_date"])
        self._to_date_var.set(state["to_date"])
        self._projects_var.set(state["projects"])
        self._issuetypes_var.set(state["issuetypes"])
        self._terminology_var.set(state["terminology"])
        self._ct_method_var.set(state["ct_method"])

        for mid, var in self._metric_vars.items():
            if mid in state["metrics"]:
                var.set(bool(state["metrics"][mid]))

        # Warn if stored file paths have gone missing
        for key in ("issue_times", "cfd"):
            p = state[key]
            if p and not Path(p).is_file():
                self._log(self._tr("err_not_found").format(p))

        # Reload filter options if IssueTimes path is set and valid
        it_path = state["issue_times"]
        if it_path and Path(it_path).is_file():
            self._load_filter_options_async(Path(it_path))

        # Apply language last (triggers _apply_language which rebuilds menu labels)
        self._lang_var.set(state["language"])

        self._log(self._tr("log_tpl_loaded").format(Path(path).name))

    # -------------------------------------------------------------------------
    # File pickers
    # -------------------------------------------------------------------------

    def _pick_issue_times(self) -> None:
        """Open a file dialog to select IssueTimes.xlsx; loads filter options and runs stage check."""
        path = filedialog.askopenfilename(
            title=self._tr("dlg_issue_times"),
            filetypes=[("Excel-Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
        )
        if path:
            self._issue_times_var.set(path)
            self._load_filter_options_async(Path(path))
            cfd = self._cfd_var.get().strip()
            if cfd:
                self._check_consistency_async(Path(path), Path(cfd))

    def _pick_cfd(self) -> None:
        """Open a file dialog to select CFD.xlsx; runs stage consistency check if IssueTimes is set."""
        path = filedialog.askopenfilename(
            title=self._tr("dlg_cfd"),
            filetypes=[("Excel-Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
        )
        if path:
            self._cfd_var.set(path)
            it = self._issue_times_var.get().strip()
            if it:
                self._check_consistency_async(Path(it), Path(path))

    def _load_filter_options_async(self, path: Path) -> None:
        """
        Load available projects and issue types from the IssueTimes file in a background thread.

        Updates self._available_projects and self._available_issuetypes when done.

        Args:
            path: Path to the IssueTimes.xlsx file.
        """
        def worker() -> None:
            try:
                projects, issuetypes = _read_available_filters(path)
                self._available_projects = projects
                self._available_issuetypes = issuetypes
                self._log(
                    f"  {len(projects)} Projekte, {len(issuetypes)} Issuetypen geladen."
                )
            except Exception as exc:
                self._log(self._tr("log_error").format(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _check_consistency_async(self, it_path: Path, cfd_path: Path) -> None:
        """
        Check stage name consistency between IssueTimes and CFD in a background thread.

        Logs the result (OK or list of mismatches) to the log area.

        Args:
            it_path:  Path to IssueTimes.xlsx.
            cfd_path: Path to CFD.xlsx.
        """
        def worker() -> None:
            try:
                only_it, only_cfd = _check_stage_consistency(it_path, cfd_path)
                if not only_it and not only_cfd:
                    self._log(self._tr("log_check_ok"))
                else:
                    if only_it:
                        self._log(self._tr("log_check_miss_cfd").format(", ".join(only_it)))
                    if only_cfd:
                        self._log(self._tr("log_check_miss_it").format(", ".join(only_cfd)))
            except Exception as exc:
                self._log(self._tr("log_error").format(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _open_multiselect(
        self,
        target_var: tk.StringVar,
        available: list[str],
        title_key: str,
    ) -> None:
        """
        Open a modal popup with checkboxes for selecting multiple filter values.

        Pre-checks items already present in target_var (comma-separated).
        On confirmation, writes the selected values back as comma-separated string.

        Args:
            target_var:  StringVar of the filter entry to read from and write to.
            available:   List of available values to show as checkboxes.
            title_key:   Translation key for the popup window title.
        """
        top = tk.Toplevel(self)
        top.title(self._tr(title_key))
        top.resizable(False, True)
        top.grab_set()

        current = {v.strip() for v in target_var.get().split(",") if v.strip()}

        # Scrollable frame for checkboxes
        frame = tk.Frame(top)
        frame.pack(fill="both", expand=True, padx=10, pady=(10, 4))

        canvas = tk.Canvas(frame, width=260, height=min(300, max(60, len(available) * 24)))
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas)

        inner.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        check_vars: dict[str, tk.BooleanVar] = {}
        for value in available:
            var = tk.BooleanVar(value=value in current)
            check_vars[value] = var
            tk.Checkbutton(inner, text=value, variable=var, anchor="w").pack(
                fill="x", padx=4, pady=1
            )

        if not available:
            tk.Label(inner, text="–", fg="gray").pack(padx=8, pady=4)

        # Select-all / clear buttons
        ctrl_frame = tk.Frame(top)
        ctrl_frame.pack(fill="x", padx=10, pady=2)
        ttk.Button(ctrl_frame, text=self._tr("btn_all"), width=6,
                   command=lambda: [v.set(True) for v in check_vars.values()]
                   ).pack(side="left", padx=2)
        ttk.Button(ctrl_frame, text=self._tr("btn_none"), width=6,
                   command=lambda: [v.set(False) for v in check_vars.values()]
                   ).pack(side="left", padx=2)

        def _confirm() -> None:
            selected = [v for v, var in check_vars.items() if var.get()]
            target_var.set(", ".join(selected))
            top.destroy()

        ttk.Button(top, text=self._tr("btn_ok"), command=_confirm).pack(pady=(4, 10))
        self.wait_window(top)

    def _pick_date(self, target_var: tk.StringVar) -> None:
        """
        Open a modal calendar popup and write the selected date to target_var.

        Pre-fills the calendar with the currently entered date if valid,
        otherwise uses today as starting point.

        Args:
            target_var: StringVar of the date entry field to update.
        """
        try:
            current = date.fromisoformat(target_var.get().strip())
        except ValueError:
            current = date.today()

        top = tk.Toplevel(self)
        top.title(self._tr("dlg_pick_date"))
        top.resizable(False, False)
        top.grab_set()

        cal = Calendar(
            top,
            selectmode="day",
            year=current.year,
            month=current.month,
            day=current.day,
            date_pattern="yyyy-mm-dd",
        )
        cal.pack(padx=10, pady=10)

        def _confirm() -> None:
            target_var.set(cal.get_date())
            top.destroy()

        ttk.Button(top, text=self._tr("btn_ok"), command=_confirm).pack(pady=(0, 10))
        self.wait_window(top)

    # -------------------------------------------------------------------------
    # Metric selection helpers
    # -------------------------------------------------------------------------

    def _set_last_365_days(self) -> None:
        """Set the From/To date fields to cover the last 365 days ending today."""
        from_d, to_d = _default_date_range()
        self._from_date_var.set(str(from_d))
        self._to_date_var.set(str(to_d))

    def _select_all_metrics(self) -> None:
        """Set all metric checkboxes to checked."""
        for var in self._metric_vars.values():
            var.set(True)

    def _deselect_all_metrics(self) -> None:
        """Set all metric checkboxes to unchecked."""
        for var in self._metric_vars.values():
            var.set(False)

    # -------------------------------------------------------------------------
    # Input parsing
    # -------------------------------------------------------------------------

    def _read_inputs(self) -> dict | None:
        """
        Validate and read all user inputs from the UI.

        Returns:
            Dict with keys: issue_times, cfd, from_date, to_date, projects,
            issuetypes, terminology, ct_method, metrics.
            Returns None if validation fails.
        """
        issue_times_str = self._issue_times_var.get().strip()
        if not issue_times_str:
            self._log(self._tr("err_no_file"))
            return None
        issue_times = Path(issue_times_str)
        if not issue_times.is_file():
            self._log(self._tr("err_not_found").format(issue_times))
            return None

        cfd_str = self._cfd_var.get().strip()
        cfd = Path(cfd_str) if cfd_str else None

        try:
            from_date = _parse_date_safe(self._from_date_var.get())
        except ValueError:
            self._log(self._tr("err_from_date").format(self._from_date_var.get()))
            return None

        try:
            to_date = _parse_date_safe(self._to_date_var.get())
        except ValueError:
            self._log(self._tr("err_to_date").format(self._to_date_var.get()))
            return None

        projects = _split_csv(self._projects_var.get())
        issuetypes = _split_csv(self._issuetypes_var.get())
        terminology = self._terminology_var.get()
        ct_method = self._ct_method_var.get()

        selected = [mid for mid, var in self._metric_vars.items() if var.get()]
        metrics = selected if selected else None

        return dict(
            issue_times=issue_times,
            cfd=cfd,
            from_date=from_date,
            to_date=to_date,
            projects=projects,
            issuetypes=issuetypes,
            terminology=terminology,
            ct_method=ct_method,
            metrics=metrics,
        )

    # -------------------------------------------------------------------------
    # Actions
    # -------------------------------------------------------------------------

    def _run_show(self) -> None:
        """Validate inputs, compute metrics, and display charts in the browser."""
        inputs = self._read_inputs()
        if inputs is None:
            return
        self._log(self._tr("log_started"))
        self._show_in_browser_from_inputs(inputs)

    def _show_in_browser_from_inputs(self, inputs: dict) -> None:
        """
        Compute metrics in a background thread and open results in the browser.

        Args:
            inputs: Dict of validated pipeline inputs from _read_inputs().
        """
        self._set_buttons_enabled(False)

        def worker() -> None:
            from .filters import FilterConfig, apply_filters
            from .loader import load_report_data
            from .metrics import all_metrics as _all_metrics
            from .metrics import get_metric

            try:
                data = load_report_data(inputs["issue_times"], inputs["cfd"])
                cfg = FilterConfig(
                    from_date=inputs["from_date"],
                    to_date=inputs["to_date"],
                    projects=inputs["projects"] or [],
                    issuetypes=inputs["issuetypes"] or [],
                )
                data = apply_filters(data, cfg)

                if inputs["metrics"]:
                    plugins = []
                    for mid in inputs["metrics"]:
                        try:
                            plugins.append(get_metric(mid))
                        except KeyError:
                            self._log(self._tr("log_unk_metric").format(mid))
                else:
                    plugins = _all_metrics()

                for plugin in plugins:
                    if isinstance(plugin, FlowTimeMetric):
                        plugin.ct_method = inputs.get("ct_method", CT_METHOD_A)

                all_figures = []
                all_results = []
                section_breaks: dict[int, str] = {}
                for plugin in plugins:
                    result = plugin.compute(data, inputs["terminology"])
                    all_results.append(result)
                    for w in result.warnings:
                        self._log(f"  WARNING: {w}")
                    figs = plugin.render(result, inputs["terminology"])
                    if figs:
                        section_breaks[len(all_figures)] = term(
                            plugin.metric_id, inputs["terminology"]
                        )
                    all_figures.extend(figs)

                # Log zero-day issues to the log window (browser mode)
                seen_zd: set[str] = set()
                zero_day_keys: list[str] = []
                for res in all_results:
                    for rec in res.stats.get("zero_day_records", []):
                        if rec.key not in seen_zd:
                            seen_zd.add(rec.key)
                            zero_day_keys.append(rec.key)
                if zero_day_keys:
                    self._log(self._tr("log_zero_day_log").format(
                        len(zero_day_keys), " | ".join(zero_day_keys)
                    ))

                if not all_figures:
                    self._log(self._tr("log_no_figs"))
                    return

                html = _build_combined_html(all_figures, section_breaks)
                with tempfile.NamedTemporaryFile(
                    mode="w", suffix=".html", delete=False, encoding="utf-8"
                ) as f:
                    f.write(html)
                    tmp_path = f.name

                self._log(self._tr("log_figs_opened").format(len(all_figures)))
                webbrowser.open(f"file:///{tmp_path.replace(chr(92), '/')}")

            except Exception as exc:
                self._log(self._tr("log_error").format(exc))
            finally:
                self.after(0, lambda: self._set_buttons_enabled(True))
                self.after(0, lambda: self._log(self._tr("log_done")))

        threading.Thread(target=worker, daemon=True).start()

    def _run_export_pdf(self) -> None:
        """Validate inputs, ask for save path, compute metrics, and export PDF."""
        inputs = self._read_inputs()
        if inputs is None:
            return

        out_path = filedialog.asksaveasfilename(
            title=self._tr("dlg_pdf"),
            defaultextension=".pdf",
            filetypes=[("PDF-Dateien", "*.pdf"), ("Alle Dateien", "*.*")],
        )
        if not out_path:
            return

        self._set_buttons_enabled(False)
        self._log(self._tr("log_pdf_started"))

        def worker() -> None:
            try:
                run_reports(
                    issue_times=inputs["issue_times"],
                    cfd=inputs["cfd"],
                    metrics=inputs["metrics"],
                    from_date=inputs["from_date"],
                    to_date=inputs["to_date"],
                    projects=inputs["projects"],
                    issuetypes=inputs["issuetypes"],
                    terminology=inputs["terminology"],
                    ct_method=inputs["ct_method"],
                    output_pdf=Path(out_path),
                    log=self._log,
                )
                self._log(self._tr("log_done"))
            except Exception as exc:
                self._log(self._tr("log_error").format(exc))
            finally:
                self.after(0, lambda: self._set_buttons_enabled(True))

        threading.Thread(target=worker, daemon=True).start()

    # -------------------------------------------------------------------------
    # Helpers
    # -------------------------------------------------------------------------

    def _log(self, msg: str) -> None:
        """
        Append a message to the log area in a thread-safe manner.

        Args:
            msg: Text to append.
        """
        def _append() -> None:
            self._log_area.configure(state="normal")
            self._log_area.insert("end", msg + "\n")
            self._log_area.see("end")
            self._log_area.configure(state="disabled")

        self.after(0, _append)

    def _set_buttons_enabled(self, enabled: bool) -> None:
        """
        Enable or disable the action buttons.

        Args:
            enabled: True to enable, False to disable.
        """
        state = "normal" if enabled else "disabled"
        self._show_btn.configure(state=state)
        self._pdf_btn.configure(state=state)


def main() -> None:
    """
    Launch the build_reports GUI.

    Entry point called by __main__.py when no CLI arguments are provided.
    """
    BuildReportsApp().mainloop()


if __name__ == "__main__":
    main()
